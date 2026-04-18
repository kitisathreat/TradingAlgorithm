import pandas as pd
import requests
import json
from datetime import datetime, timedelta
import time
from typing import Dict, List, Optional, Tuple
import warnings
import os
import sys
from difflib import SequenceMatcher
import re
import numpy as np
warnings.filterwarnings('ignore')

# Optional performance dependencies — the module degrades gracefully if absent.
try:
    from joblib import Parallel, delayed  # type: ignore
    JOBLIB_AVAILABLE = True
except ImportError:
    JOBLIB_AVAILABLE = False
    Parallel = None  # type: ignore
    delayed = None   # type: ignore

try:
    from rapidfuzz import fuzz, process as rf_process  # type: ignore
    RAPIDFUZZ_AVAILABLE = True
except ImportError:
    RAPIDFUZZ_AVAILABLE = False

class SECFileSourcer:
    """
    A comprehensive class for sourcing SEC filings and creating financial models.
    
    This class provides methods to:
    1. Find and download SEC 10-K and 10-Q filings
    2. Create traditional financial three-statement models
    3. Generate sensitivity analysis with operating leverage impacts
    4. Create KPI summary sheets
    """
    
    # Class-level cache for ticker-to-CIK mapping
    _ticker_cik_cache = None
    _cache_last_updated = None
    _cache_expiry_seconds = 24 * 3600  # 24 hours
    _fuzzy_match_cache = {}
    
    def __init__(self):
        """Initialize the SEC File Sourcer with base URLs and headers."""
        self.base_url = "https://data.sec.gov"
        self.sec_ticker_url = "https://www.sec.gov/files/company_tickers.json"
        self.headers = {
            'User-Agent': 'ModelMaker/1.0 (kit.kumar@gmail.com)',
            'Accept': 'application/json',
            'Accept-Encoding': 'gzip, deflate'
        }
        self.session = requests.Session()
        self.session.headers.update(self.headers)
        
        # Add rate limiting - SEC requires delays between requests
        self.last_request_time = 0
        self.min_request_interval = 0.1  # 100ms minimum between requests
        
    def _rate_limit(self):
        """Ensure minimum time between API requests to comply with SEC rate limits."""
        current_time = time.time()
        time_since_last = current_time - self.last_request_time
        if time_since_last < self.min_request_interval:
            time.sleep(self.min_request_interval - time_since_last)
        self.last_request_time = time.time()

    @classmethod
    def _is_cache_valid(cls):
        if cls._ticker_cik_cache is None or cls._cache_last_updated is None:
            return False
        return (time.time() - cls._cache_last_updated) < cls._cache_expiry_seconds

    @classmethod
    def _load_ticker_cik_cache(cls, session, url):
        """Load the ticker-to-CIK mapping from the SEC and cache it."""
        try:
            response = session.get(url)
            if response.status_code != 200:
                print(f"Error fetching company tickers: {response.status_code}")
                return None
            data = response.json()
            # The SEC file is now a dictionary with numeric string keys
            cache = {}
            for key, entry in data.items():
                if isinstance(entry, dict) and 'ticker' in entry and 'cik_str' in entry:
                    cache[entry['ticker'].upper()] = str(entry['cik_str']).zfill(10)
            cls._ticker_cik_cache = cache
            cls._cache_last_updated = time.time()
            return cache
        except Exception as e:
            print(f"Error loading ticker-to-CIK cache: {e}")
            return None

    def get_cik_from_ticker(self, ticker: str, force_refresh: bool = False) -> Optional[str]:
        """
        Convert a stock ticker symbol to its corresponding CIK (Central Index Key) number.
        Uses a cached mapping for efficiency.
        Args:
            ticker (str): Stock ticker symbol (e.g., 'AAPL', 'MSFT')
            force_refresh (bool): If True, refresh the cache from the SEC
        Returns:
            Optional[str]: CIK number as a string (10-digit zero-padded), or None if not found
        """
        ticker_upper = ticker.upper()
        # Check cache
        if force_refresh or not self._is_cache_valid():
            self._rate_limit()
            cache = self._load_ticker_cik_cache(self.session, self.sec_ticker_url)
            if cache is None:
                print("Could not fetch ticker-to-CIK mapping from SEC.")
                return None
        else:
            cache = self._ticker_cik_cache
        cik = cache.get(ticker_upper)
        if cik:
            return cik
        print(f"Ticker '{ticker}' not found in SEC database.")
        return None
    
    def find_sec_filings(self, ticker: str, filing_types: List[str] = ['10-K', '10-Q']) -> pd.DataFrame:
        """
        Find SEC 10-K and 10-Q filings for a given stock ticker, sorted by date.
        
        Args:
            ticker (str): Stock ticker symbol
            filing_types (List[str]): Types of filings to search for (default: ['10-K', '10-Q'])
            
        Returns:
            pd.DataFrame: DataFrame containing filing information sorted by date
        """
        try:
            # Convert ticker to CIK
            cik = self.get_cik_from_ticker(ticker)
            if not cik:
                print(f"Could not find CIK for ticker: {ticker}")
                return pd.DataFrame()
            
            # Get company submissions
            submissions_url = f"{self.base_url}/submissions/CIK{cik}.json"
            response = self.session.get(submissions_url)
            
            if response.status_code != 200:
                print(f"Error fetching submissions for {ticker} (CIK: {cik}): {response.status_code}")
                return pd.DataFrame()
            
            submissions_data = response.json()
            filings = submissions_data.get('filings', {}).get('recent', {})
            
            # Create DataFrame from filings
            filing_data = []
            for i in range(len(filings.get('form', []))):
                form = filings['form'][i]
                if form in filing_types:
                    filing_data.append({
                        'form': form,
                        'filingDate': filings['filingDate'][i],
                        'accessionNumber': filings['accessionNumber'][i],
                        'primaryDocument': filings['primaryDocument'][i],
                        'description': filings.get('description', [''])[i] if i < len(filings.get('description', [])) else ''
                    })
            
            df = pd.DataFrame(filing_data)
            if not df.empty:
                df['filingDate'] = pd.to_datetime(df['filingDate'])
                df = df.sort_values('filingDate', ascending=False)
                
            return df
            
        except Exception as e:
            print(f"Error in find_sec_filings: {str(e)}")
            return pd.DataFrame()
    
    def create_financial_model(self, ticker: str, quarters: int = 8, progress_callback=None, enhanced_fuzzy_matching: bool = True) -> Dict[str, pd.DataFrame]:
        """
        Create a traditional financial three-statement model with annual and quarterly views.
        
        Args:
            ticker (str): Stock ticker symbol
            quarters (int): Number of quarters of data to retrieve (default: 8, which is 2 years)
                          This will determine how many 10-K and 10-Q filings to process.
                          For example: 4 quarters = 1 year, 8 quarters = 2 years, 12 quarters = 3 years
            progress_callback (callable): Optional callback function for progress updates
            enhanced_fuzzy_matching (bool): Whether to include non-GAAP to GAAP mapping (default: True)
            
        Returns:
            Dict[str, pd.DataFrame]: Dictionary containing financial model dataframes
        """
        def progress(msg):
            if progress_callback:
                progress_callback(msg)
            else:
                print(msg)
        
        try:
            # Store the quarters parameter for use in data filtering
            self.current_quarters = quarters
            
            # Initialize the financial model structure
            financial_model = {
                'annual_income_statement': pd.DataFrame(),
                'annual_balance_sheet': pd.DataFrame(),
                'annual_cash_flow': pd.DataFrame(),
                'quarterly_income_statement': pd.DataFrame(),
                'quarterly_balance_sheet': pd.DataFrame(),
                'quarterly_cash_flow': pd.DataFrame()
            }
            
            # Convert ticker to CIK
            progress("    • Converting ticker to CIK number...")
            cik = self.get_cik_from_ticker(ticker)
            if not cik:
                progress(f"    ✗ Could not find CIK for ticker: {ticker}")
                return financial_model
            
            progress(f"    ✓ Found CIK: {cik}")
            
            # Remove leading zeros from CIK for new endpoint
            cik_no_zeros = str(int(cik))
            # Get company facts from new endpoint
            progress("    • Fetching company facts from SEC API...")
            company_facts_url = f"https://data.sec.gov/api/xbrl/companyfacts/CIK{cik_no_zeros}.json"
            response = self.session.get(company_facts_url)
            
            if response.status_code == 200:
                progress("    ✓ Successfully retrieved company facts from SEC API")
                facts_data = response.json()
                facts = facts_data.get('facts', {})
                # Define key financial metrics for each statement
                income_statement_metrics = {
                    'Revenues': 'us-gaap:Revenues',
                    'CostOfRevenue': 'us-gaap:CostOfRevenue',
                    'GrossProfit': 'us-gaap:GrossProfit',
                    'OperatingExpenses': 'us-gaap:OperatingExpenses',
                    'OperatingIncomeLoss': 'us-gaap:OperatingIncomeLoss',
                    'NetIncomeLoss': 'us-gaap:NetIncomeLoss',
                    'EarningsPerShareBasic': 'us-gaap:EarningsPerShareBasic',
                    'EarningsPerShareDiluted': 'us-gaap:EarningsPerShareDiluted'
                }
                balance_sheet_metrics = {
                    'CashAndCashEquivalents': 'us-gaap:CashAndCashEquivalentsAtCarryingValue',
                    'TotalAssets': 'us-gaap:Assets',
                    'TotalCurrentAssets': 'us-gaap:AssetsCurrent',
                    'TotalLiabilities': 'us-gaap:Liabilities',
                    'TotalCurrentLiabilities': 'us-gaap:LiabilitiesCurrent',
                    'TotalStockholdersEquity': 'us-gaap:StockholdersEquity',
                    'TotalDebt': 'us-gaap:LongTermDebtNoncurrent'
                }
                cash_flow_metrics = {
                    'NetCashProvidedByUsedInOperatingActivities': 'us-gaap:NetCashProvidedByUsedInOperatingActivities',
                    'NetCashProvidedByUsedInInvestingActivities': 'us-gaap:NetCashProvidedByUsedInInvestingActivities',
                    'NetCashProvidedByUsedInFinancingActivities': 'us-gaap:NetCashProvidedByUsedInFinancingActivities',
                    'CapitalExpenditures': 'us-gaap:PaymentsToAcquirePropertyPlantAndEquipment',
                    'DividendsPaid': 'us-gaap:PaymentsOfDividends'
                }
                
                progress("    • Extracting financial data from API response...")
                for period in ['annual', 'quarterly']:
                    for statement, metrics in [('income_statement', income_statement_metrics),
                                             ('balance_sheet', balance_sheet_metrics),
                                             ('cash_flow', cash_flow_metrics)]:
                        df = self._extract_financial_data(facts, metrics, period)
                        financial_model[f'{period}_{statement}'] = df
                
                progress("    ✓ Financial model created from SEC API data")
                # Enforce required structure on income statements before returning
                financial_model['annual_income_statement'] = self._enforce_income_statement_structure(financial_model.get('annual_income_statement', pd.DataFrame()))
                financial_model['quarterly_income_statement'] = self._enforce_income_statement_structure(financial_model.get('quarterly_income_statement', pd.DataFrame()))
                return financial_model
            else:
                progress(f"    ⚠ SEC API not available (status: {response.status_code}), falling back to XBRL parsing...")
                # Fallback: Try to find and parse XBRL XML instance document using Arelle
                try:
                    progress("    • Finding SEC filings for XBRL extraction...")
                    filings_df = self.find_sec_filings(ticker)
                    if filings_df.empty:
                        progress(f"    ✗ No filings found for {ticker}")
                        return financial_model
                    
                    # Calculate how many filings we need based on quarters parameter
                    years_needed = max(1, (quarters + 3) // 4)  # Round up to get full years needed
                    k_filings_needed = years_needed  # One 10-K per year
                    q_filings_needed = min(quarters, 20)  # Limit 10-Q filings to prevent excessive processing
                    
                    progress(f"    • Requested {quarters} quarters of data ({years_needed} years)")
                    progress(f"    • Will process up to {k_filings_needed} 10-K filings and {q_filings_needed} 10-Q filings")
                    
                    # Separate 10-K and 10-Q filings
                    k_filings = filings_df[filings_df['form'] == '10-K'].head(k_filings_needed)
                    q_filings = filings_df[filings_df['form'] == '10-Q'].head(q_filings_needed)
                    
                    progress(f"    ✓ Found {len(k_filings)} 10-K filings and {len(q_filings)} 10-Q filings")
                    progress("    • Starting comprehensive XBRL data extraction...")
                    
                    # Initialize comprehensive financial model
                    comprehensive_facts = {}
                    annual_data = {}
                    quarterly_data = {}
                    
                    # Process 10-K filings first (primary source for annual data)
                    progress(f"    • Processing {len(k_filings)} 10-K filings (annual data)...")
                    for i, (idx, row) in enumerate(k_filings.iterrows(), 1):
                        progress(f"      [{i}/{len(k_filings)}] Processing 10-K: {row['filingDate']}")
                        k_facts = self._extract_xbrl_from_filing(row, cik, progress_callback)
                        if k_facts:
                            # Store annual data from 10-K
                            for concept, data in k_facts.items():
                                if concept not in annual_data:
                                    annual_data[concept] = []
                                annual_data[concept].extend(data)
                            comprehensive_facts.update(k_facts)
                            progress(f"        ✓ Extracted {len(k_facts)} concepts")
                        else:
                            progress(f"        ✗ No data extracted")
                    
                    # Process 10-Q filings (supplementary quarterly data)
                    progress(f"    • Processing {len(q_filings)} 10-Q filings (quarterly data)...")
                    for i, (idx, row) in enumerate(q_filings.iterrows(), 1):
                        progress(f"      [{i}/{len(q_filings)}] Processing 10-Q: {row['filingDate']}")
                        q_facts = self._extract_xbrl_from_filing(row, cik, progress_callback)
                        if q_facts:
                            # Store quarterly data from 10-Q
                            for concept, data in q_facts.items():
                                if concept not in quarterly_data:
                                    quarterly_data[concept] = []
                                quarterly_data[concept].extend(data)
                            comprehensive_facts.update(q_facts)
                            progress(f"        ✓ Extracted {len(q_facts)} concepts")
                        else:
                            progress(f"        ✗ No data extracted")
                    
                    if not comprehensive_facts:
                        progress("    ✗ No XBRL data found in any filings")
                        return financial_model
                    
                    progress(f"    ✓ Comprehensive data extracted: {len(comprehensive_facts)} unique concepts")
                    progress(f"    • Annual data: {sum(len(data) for data in annual_data.values())} data points")
                    progress(f"    • Quarterly data: {sum(len(data) for data in quarterly_data.values())} data points")
                    
                    # Run discrepancy checks between annual and quarterly data
                    progress("    • Running data consistency checks...")
                    self._run_discrepancy_checks(annual_data, quarterly_data)
                    
                    # Create financial model using comprehensive data
                    progress("    • Creating comprehensive financial model...")
                    financial_model = self._create_model_from_comprehensive_data(comprehensive_facts, annual_data, quarterly_data, enhanced_fuzzy_matching, progress_callback)
                    
                    progress("    ✓ Comprehensive financial model created successfully!")
                    # Enforce required structure on income statements before returning
                    financial_model['annual_income_statement'] = self._enforce_income_statement_structure(financial_model.get('annual_income_statement', pd.DataFrame()))
                    financial_model['quarterly_income_statement'] = self._enforce_income_statement_structure(financial_model.get('quarterly_income_statement', pd.DataFrame()))
                    return financial_model
                    
                except Exception as e:
                    progress(f"    ✗ Error in comprehensive XBRL processing: {str(e)}")
                    import traceback
                    traceback.print_exc()
                    return financial_model
        except Exception as e:
            progress(f"    ✗ Error creating financial model: {str(e)}")
            return financial_model

    def _extract_financial_data(self, facts: Dict, metrics: Dict, period: str) -> pd.DataFrame:
        """
        Extract financial data from SEC facts for a given period and set of metrics.
        Now includes validation to filter out non-financial data.
        
        Args:
            facts (Dict): SEC facts data
            metrics (Dict): Dictionary of metric names and their SEC tags
            period (str): 'annual' or 'quarterly'
            
        Returns:
            pd.DataFrame: DataFrame with validated financial data
        """
        data = {}
        validation_stats = {
            'total_points': 0,
            'valid_points': 0,
            'rejected_points': 0
        }
        
        for metric_name, sec_tag in metrics.items():
            if sec_tag in facts:
                metric_data = facts[sec_tag]
                units = metric_data.get('units', {})
                
                # Find the appropriate unit (USD is most common)
                unit_key = None
                for key in units.keys():
                    if 'USD' in key or key == 'USD':
                        unit_key = key
                        break
                
                if unit_key:
                    periods = units[unit_key]
                    
                    # Filter by period
                    filtered_periods = []
                    for period_data in periods:
                        if period == 'annual' and period_data.get('form') in ['10-K', '10-K/A']:
                            filtered_periods.append(period_data)
                        elif period == 'quarterly' and period_data.get('form') in ['10-Q', '10-Q/A']:
                            filtered_periods.append(period_data)
                    
                    # Sort by end date (most recent first), then deduplicate by
                    # period end. 10-K/A and 10-Q/A amendments share an `end`
                    # with the original filing; keeping the most recent entry
                    # prevents the quarterly limit from silently undercounting
                    # the number of distinct reporting periods returned.
                    filtered_periods.sort(key=lambda x: (x.get('end', ''), x.get('filed', '')), reverse=True)
                    seen_ends = set()
                    deduped_periods = []
                    for p in filtered_periods:
                        end = p.get('end', '')
                        if end and end not in seen_ends:
                            seen_ends.add(end)
                            deduped_periods.append(p)
                    filtered_periods = deduped_periods

                    # Take the most recent periods based on the quarters parameter
                    if period == 'annual':
                        years_needed = max(1, (self.current_quarters + 3) // 4)
                        recent_periods = filtered_periods[:years_needed]
                    else:
                        recent_periods = filtered_periods[:self.current_quarters]
                    
                    for period_data in recent_periods:
                        end_date = period_data.get('end', '')
                        value = period_data.get('val', 0)
                        
                        validation_stats['total_points'] += 1
                        
                        # Validate the data point
                        if self._validate_financial_data(metric_name, value, unit_key):
                            if end_date not in data:
                                data[end_date] = {}
                            data[end_date][metric_name] = value
                            validation_stats['valid_points'] += 1
                        else:
                            validation_stats['rejected_points'] += 1
        
        # Log validation statistics for SEC API path
        if hasattr(self, 'progress_callback') and self.progress_callback:
            self.progress_callback(f"    • SEC API validation: {validation_stats['valid_points']}/{validation_stats['total_points']} points passed validation")
            if validation_stats['rejected_points'] > 0:
                self.progress_callback(f"    • Rejected {validation_stats['rejected_points']} non-financial data points from SEC API")
        
        # Convert to DataFrame
        if data:
            df = pd.DataFrame.from_dict(data, orient='index')
            df.index = pd.to_datetime(df.index)
            df = df.sort_index()
            return df
        else:
            return pd.DataFrame()
    
    def create_sensitivity_model(self, financial_model: Dict[str, pd.DataFrame], 
                               ticker: str, quarters: int = 8, progress_callback=None) -> Dict[str, pd.DataFrame]:
        """
        Create a sensitivity analysis model with operating leverage impacts.
        
        Args:
            financial_model (Dict[str, pd.DataFrame]): Base financial model
            ticker (str): Stock ticker symbol
            quarters (int): Number of quarters used in the financial model (for reference)
            progress_callback (callable): Optional callback function for progress updates
            
        Returns:
            Dict[str, pd.DataFrame]: Dictionary containing sensitivity analysis dataframes
        """
        def progress(msg):
            if progress_callback:
                progress_callback(msg)
            else:
                print(msg)
        
        try:
            sensitivity_model = {
                'case_summary': pd.DataFrame(),
                'financial_model': pd.DataFrame(),
                'kpi_summary': pd.DataFrame()
            }
            
            # Get the most recent annual income statement for base calculations
            annual_income = financial_model.get('annual_income_statement', pd.DataFrame())
            
            if annual_income.empty:
                progress("    ✗ No annual income statement data available for sensitivity analysis")
                return sensitivity_model
            
            progress("    • Creating operating leverage scenarios...")
            # Create case summary sheet with operating leverage scenarios
            case_summary = self._create_case_summary(annual_income)
            sensitivity_model['case_summary'] = case_summary
            progress("    ✓ Operating leverage scenarios created")
            
            progress("    • Creating enhanced financial model with historical and forecasted data...")
            # Create enhanced financial model with historical and forecasted data
            enhanced_model = self._create_enhanced_financial_model(financial_model)
            sensitivity_model['financial_model'] = enhanced_model
            progress("    ✓ Enhanced financial model created")
            
            progress("    • Generating KPI summary sheet...")
            # Create KPI summary sheet
            kpi_summary = self._create_kpi_summary(financial_model)
            sensitivity_model['kpi_summary'] = kpi_summary
            progress("    ✓ KPI summary sheet created")
            
            return sensitivity_model
            
        except Exception as e:
            progress(f"    ✗ Error in create_sensitivity_model: {str(e)}")
            return {
                'case_summary': pd.DataFrame(),
                'financial_model': pd.DataFrame(),
                'kpi_summary': pd.DataFrame()
            }
    
    def _create_case_summary(self, annual_income: pd.DataFrame) -> pd.DataFrame:
        """
        Create case summary sheet showing operating leverage impacts.
        
        Args:
            annual_income (pd.DataFrame): Annual income statement data
            
        Returns:
            pd.DataFrame: Case summary with sensitivity scenarios
        """
        if annual_income.empty:
            return pd.DataFrame()
        
        # Get the most recent year's data
        latest_year = annual_income.index[-1]
        base_data = annual_income.loc[latest_year]
        
        # Define sensitivity scenarios
        scenarios = {
            'Base Case': 1.0,
            'Optimistic (+20%)': 1.2,
            'Pessimistic (-20%)': 0.8,
            'High Growth (+50%)': 1.5,
            'Recession (-30%)': 0.7
        }
        
        case_summary_data = []
        
        for scenario_name, revenue_multiplier in scenarios.items():
            # Calculate revenue impact
            base_revenue = base_data.get('Revenues', 0)
            new_revenue = base_revenue * revenue_multiplier
            
            # Calculate operating leverage impact
            # Assume fixed costs remain constant and variable costs scale with revenue
            base_cogs = base_data.get('CostOfRevenue', 0)
            base_opex = base_data.get('OperatingExpenses', 0)
            
            # Assume 70% of COGS is variable, 30% of OpEx is variable
            variable_cogs_ratio = 0.7
            variable_opex_ratio = 0.3
            
            new_cogs = (base_cogs * variable_cogs_ratio * revenue_multiplier + 
                       base_cogs * (1 - variable_cogs_ratio))
            new_opex = (base_opex * variable_opex_ratio * revenue_multiplier + 
                       base_opex * (1 - variable_opex_ratio))
            
            new_gross_profit = new_revenue - new_cogs
            new_operating_income = new_gross_profit - new_opex
            
            # Calculate operating leverage
            revenue_change = (new_revenue - base_revenue) / base_revenue if base_revenue != 0 else 0
            operating_income_change = (new_operating_income - base_data.get('OperatingIncomeLoss', 0)) / base_data.get('OperatingIncomeLoss', 0) if base_data.get('OperatingIncomeLoss', 0) != 0 else 0
            
            operating_leverage = operating_income_change / revenue_change if revenue_change != 0 else 0
            
            case_summary_data.append({
                'Scenario': scenario_name,
                'Revenue': new_revenue,
                'COGS': new_cogs,
                'Gross Profit': new_gross_profit,
                'Operating Expenses': new_opex,
                'Operating Income': new_operating_income,
                'Revenue Change %': revenue_change * 100,
                'Operating Income Change %': operating_income_change * 100,
                'Operating Leverage': operating_leverage
            })
        
        return pd.DataFrame(case_summary_data)
    
    def _create_enhanced_financial_model(self, financial_model: Dict[str, pd.DataFrame]) -> pd.DataFrame:
        """
        Create enhanced financial model with historical and forecasted data.
        
        Args:
            financial_model (Dict[str, pd.DataFrame]): Base financial model
            
        Returns:
            pd.DataFrame: Enhanced financial model
        """
        # Combine all financial data into one comprehensive model
        enhanced_data = {}
        
        # Add historical data
        for statement_type in ['income_statement', 'balance_sheet', 'cash_flow']:
            for period in ['annual', 'quarterly']:
                key = f'{period}_{statement_type}'
                if key in financial_model and not financial_model[key].empty:
                    df = financial_model[key]
                    for date in df.index:
                        for column in df.columns:
                            enhanced_data[f"{period}_{statement_type}_{column}"] = {
                                'date': date,
                                'value': df.loc[date, column],
                                'type': 'historical',
                                'period': period,
                                'statement': statement_type
                            }
        
        # Add forecasted data (simple linear projection for demonstration)
        if enhanced_data:
            # Get the most recent dates for each period
            annual_dates = []
            quarterly_dates = []
            
            for key, data in enhanced_data.items():
                if data['period'] == 'annual':
                    annual_dates.append(data['date'])
                elif data['period'] == 'quarterly':
                    quarterly_dates.append(data['date'])
            
            if annual_dates:
                latest_annual = max(annual_dates)
                # Add 3 years of forecast
                for i in range(1, 4):
                    forecast_date = latest_annual + pd.DateOffset(years=i)
                    enhanced_data[f"forecast_annual_year_{i}"] = {
                        'date': forecast_date,
                        'value': 0,  # Placeholder for forecasted values
                        'type': 'forecast',
                        'period': 'annual',
                        'statement': 'projection'
                    }
            
            if quarterly_dates:
                latest_quarterly = max(quarterly_dates)
                # Add 4 quarters of forecast
                for i in range(1, 5):
                    forecast_date = latest_quarterly + pd.DateOffset(months=i*3)
                    enhanced_data[f"forecast_quarterly_q{i}"] = {
                        'date': forecast_date,
                        'value': 0,  # Placeholder for forecasted values
                        'type': 'forecast',
                        'period': 'quarterly',
                        'statement': 'projection'
                    }
        
        # Convert to DataFrame
        if enhanced_data:
            df = pd.DataFrame.from_dict(enhanced_data, orient='index')
            df = df.sort_values('date')
            return df
        else:
            return pd.DataFrame()
    
    def _create_kpi_summary(self, financial_model: Dict[str, pd.DataFrame]) -> pd.DataFrame:
        """
        Create KPI summary sheet with key financial metrics.
        
        Args:
            financial_model (Dict[str, pd.DataFrame]): Base financial model
            
        Returns:
            pd.DataFrame: KPI summary sheet
        """
        kpi_data = []
        
        # Calculate KPIs from available data
        annual_income = financial_model.get('annual_income_statement', pd.DataFrame())
        annual_balance = financial_model.get('annual_balance_sheet', pd.DataFrame())
        annual_cash_flow = financial_model.get('annual_cash_flow', pd.DataFrame())
        
        if not annual_income.empty:
            latest_income = annual_income.iloc[-1]
            
            # Profitability KPIs
            revenue = latest_income.get('Revenues', 0)
            net_income = latest_income.get('NetIncomeLoss', 0)
            operating_income = latest_income.get('OperatingIncomeLoss', 0)
            gross_profit = latest_income.get('GrossProfit', 0)
            
            if revenue != 0:
                kpi_data.extend([
                    {'KPI': 'Net Profit Margin', 'Value': (net_income / revenue) * 100, 'Unit': '%'},
                    {'KPI': 'Operating Margin', 'Value': (operating_income / revenue) * 100, 'Unit': '%'},
                    {'KPI': 'Gross Margin', 'Value': (gross_profit / revenue) * 100, 'Unit': '%'}
                ])
        
        if not annual_balance.empty:
            latest_balance = annual_balance.iloc[-1]
            
            # Efficiency KPIs
            total_assets = latest_balance.get('TotalAssets', 0)
            total_equity = latest_balance.get('TotalStockholdersEquity', 0)
            
            if total_assets != 0:
                kpi_data.append({'KPI': 'Return on Assets (ROA)', 'Value': (net_income / total_assets) * 100, 'Unit': '%'})
            
            if total_equity != 0:
                kpi_data.append({'KPI': 'Return on Equity (ROE)', 'Value': (net_income / total_equity) * 100, 'Unit': '%'})
            
            # Liquidity KPIs
            current_assets = latest_balance.get('TotalCurrentAssets', 0)
            current_liabilities = latest_balance.get('TotalCurrentLiabilities', 0)
            
            if current_liabilities != 0:
                kpi_data.append({'KPI': 'Current Ratio', 'Value': current_assets / current_liabilities, 'Unit': 'x'})
        
        if not annual_cash_flow.empty:
            latest_cash_flow = annual_cash_flow.iloc[-1]
            
            # Cash Flow KPIs
            operating_cash_flow = latest_cash_flow.get('NetCashProvidedByUsedInOperatingActivities', 0)
            capital_expenditures = abs(latest_cash_flow.get('CapitalExpenditures', 0))
            
            if capital_expenditures != 0:
                kpi_data.append({'KPI': 'Operating Cash Flow to CapEx', 'Value': operating_cash_flow / capital_expenditures, 'Unit': 'x'})
        
        # Add growth metrics if multiple years available
        if len(annual_income) >= 2:
            current_revenue = annual_income.iloc[-1].get('Revenues', 0)
            previous_revenue = annual_income.iloc[-2].get('Revenues', 0)
            
            if previous_revenue != 0:
                revenue_growth = ((current_revenue - previous_revenue) / previous_revenue) * 100
                kpi_data.append({'KPI': 'Revenue Growth (YoY)', 'Value': revenue_growth, 'Unit': '%'})
        
        return pd.DataFrame(kpi_data)
    
    def get_filing_content(self, accession_number: str, primary_document: str) -> str:
        """
        Get the content of a specific SEC filing.
        
        Args:
            accession_number (str): SEC accession number
            primary_document (str): Primary document name
            
        Returns:
            str: Filing content
        """
        try:
            # Format accession number
            accession_number = accession_number.replace('-', '')
            
            # Construct URL for filing
            filing_url = f"https://www.sec.gov/Archives/edgar/data/{accession_number}/{primary_document}"
            
            response = self.session.get(filing_url)
            
            if response.status_code == 200:
                return response.text
            else:
                print(f"Error fetching filing content: {response.status_code}")
                return ""
                
        except Exception as e:
            print(f"Error in get_filing_content: {str(e)}")
            return ""
    
    def export_to_excel_fast_preview(self, financial_model: Dict[str, pd.DataFrame], 
                                   sensitivity_model: Dict[str, pd.DataFrame], 
                                   ticker: str, filename: str, progress_callback=None) -> str:
        """
        Export financial models to Excel file with minimal formatting for fast preview generation.
        This method skips the expensive professional formatting to provide quick results.
        """
        def progress(msg):
            if progress_callback:
                progress_callback(msg)
            else:
                print(msg)
        
        import pandas as pd
        import os
        from datetime import datetime
        
        progress("    • Creating fast preview Excel file...")
        
        # Ensure the Storage directory exists
        storage_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))), 'Storage')
        os.makedirs(storage_dir, exist_ok=True)
        
        filepath = os.path.join(storage_dir, filename)
        
        progress(f"    • Writing data to Excel: {filename}")

        # Write all DataFrames to Excel using pandas (fast, no formatting)
        progress("    • Writing financial data to Excel sheets...")
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            # Annual sheet
            annual_income = financial_model.get('annual_income_statement', pd.DataFrame())
            annual_balance = financial_model.get('annual_balance_sheet', pd.DataFrame())
            annual_cash_flow = financial_model.get('annual_cash_flow', pd.DataFrame())
            if not annual_income.empty or not annual_balance.empty or not annual_cash_flow.empty:
                progress("    • Writing annual financial statements...")
                startrow = 0
                if not annual_income.empty:
                    # Remove empty columns before writing to Excel
                    cleaned_annual_income = self._remove_empty_columns(annual_income.transpose())
                    cleaned_annual_income.to_excel(writer, sheet_name='Annual Financial Statements', startrow=startrow)
                    startrow += annual_income.shape[1] + 3
                if not annual_balance.empty:
                    # Remove empty columns before writing to Excel
                    cleaned_annual_balance = self._remove_empty_columns(annual_balance.transpose())
                    cleaned_annual_balance.to_excel(writer, sheet_name='Annual Financial Statements', startrow=startrow)
                    startrow += annual_balance.shape[1] + 3
                if not annual_cash_flow.empty:
                    # Remove empty columns before writing to Excel
                    cleaned_annual_cash_flow = self._remove_empty_columns(annual_cash_flow.transpose())
                    cleaned_annual_cash_flow.to_excel(writer, sheet_name='Annual Financial Statements', startrow=startrow)
            
            # Quarterly sheet
            quarterly_income = financial_model.get('quarterly_income_statement', pd.DataFrame())
            quarterly_balance = financial_model.get('quarterly_balance_sheet', pd.DataFrame())
            quarterly_cash_flow = financial_model.get('quarterly_cash_flow', pd.DataFrame())
            if not quarterly_income.empty or not quarterly_balance.empty or not quarterly_cash_flow.empty:
                progress("    • Writing quarterly financial statements...")
                startrow = 0
                if not quarterly_income.empty:
                    # Remove empty columns before writing to Excel
                    cleaned_quarterly_income = self._remove_empty_columns(quarterly_income.transpose())
                    cleaned_quarterly_income.to_excel(writer, sheet_name='Quarterly Financial Statements', startrow=startrow)
                    startrow += quarterly_income.shape[1] + 3
                if not quarterly_balance.empty:
                    # Remove empty columns before writing to Excel
                    cleaned_quarterly_balance = self._remove_empty_columns(quarterly_balance.transpose())
                    cleaned_quarterly_balance.to_excel(writer, sheet_name='Quarterly Financial Statements', startrow=startrow)
                    startrow += quarterly_balance.shape[1] + 3
                if not quarterly_cash_flow.empty:
                    # Remove empty columns before writing to Excel
                    cleaned_quarterly_cash_flow = self._remove_empty_columns(quarterly_cash_flow.transpose())
                    cleaned_quarterly_cash_flow.to_excel(writer, sheet_name='Quarterly Financial Statements', startrow=startrow)
            
            # Sensitivity and summary sheets
            progress("    • Writing sensitivity analysis and summary sheets...")
            for sheet_name, df in sensitivity_model.items():
                if not df.empty:
                    # Remove empty columns before writing to Excel
                    cleaned_df = self._remove_empty_columns(df)
                    cleaned_df.to_excel(writer, sheet_name=sheet_name.replace('_', ' ').title())
            
            # Summary sheet
            summary_data = {
                'Metric': ['Ticker', 'Model Created', 'Data Points', 'Scenarios Analyzed'],
                'Value': [
                    ticker,
                    datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    sum(len(df) for df in financial_model.values() if not df.empty),
                    len(sensitivity_model.get('case_summary', pd.DataFrame()))
                ]
            }
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='Summary', index=False)

        progress(f"    ✓ Fast preview Excel file saved: {filepath}")
        return filepath

    def apply_excel_formatting(self, filepath: str, schmoove_mode: bool = False, progress_callback=None) -> str:
        """
        Apply professional formatting to an existing Excel file.
        This can be used to upgrade a fast preview file to full formatting.
        """
        def progress(msg):
            if progress_callback:
                progress_callback(msg)
            else:
                print(msg)
        
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        import os
        
        if not os.path.exists(filepath):
            progress(f"    ✗ File not found: {filepath}")
            return filepath
        
        progress("    • Applying professional formatting to existing Excel file...")
        
        if schmoove_mode:
            try:
                import joblib
                import psutil
                # Limit joblib to 75% of CPUs
                n_jobs = max(1, int(psutil.cpu_count(logical=True) * 0.75))
                progress(f"    • Schmoove mode enabled: Using {n_jobs} CPU cores")
            except ImportError:
                n_jobs = 1
        else:
            n_jobs = 1
        
        # Open the workbook
        wb = openpyxl.load_workbook(filepath)
        
        # Format annual and quarterly sheets
        for sheet in ['Annual Financial Statements', 'Quarterly Financial Statements']:
            if sheet in wb.sheetnames:
                ws = wb[sheet]
                
                # Find the key columns
                line_item_col = None
                section_heading_col = None
                parent_col = None
                aggregate_col = None
                
                for idx, cell in enumerate(ws[1]):
                    if cell.value == 'Line Item':
                        line_item_col = idx + 1
                    elif cell.value == 'is_section_heading':
                        section_heading_col = idx + 1
                    elif cell.value == 'parent':
                        parent_col = idx + 1
                    elif cell.value == 'is_aggregate':
                        aggregate_col = idx + 1
                
                if line_item_col is None:
                    continue
                
                # Apply formatting to each row
                for row_num in range(2, ws.max_row + 1):
                    line_item_cell = ws.cell(row=row_num, column=line_item_col)
                    line_item_value = line_item_cell.value
                    
                    if line_item_value is None:
                        continue
                    
                    # Check if this is a section heading
                    is_section_heading = False
                    if section_heading_col:
                        section_heading_cell = ws.cell(row=row_num, column=section_heading_col)
                        is_section_heading = section_heading_cell.value == True or section_heading_cell.value == 'True'
                    
                    # Check if this is an aggregator
                    is_aggregate = False
                    if aggregate_col:
                        aggregate_cell = ws.cell(row=row_num, column=aggregate_col)
                        is_aggregate = aggregate_cell.value == True or aggregate_cell.value == 'True'
                    
                    # Check if this has a parent (for indentation)
                    has_parent = False
                    if parent_col:
                        parent_cell = ws.cell(row=row_num, column=parent_col)
                        has_parent = parent_cell.value is not None and parent_cell.value != ''
                    
                    # Apply formatting based on flags
                    if is_section_heading:
                        # Bold for section headings (INCOME STATEMENT, BALANCE SHEET, CASH FLOW STATEMENT)
                        line_item_cell.font = Font(bold=True, size=12)
                        line_item_cell.alignment = Alignment(horizontal='left')
                    elif is_aggregate:
                        # Italic for aggregator line items
                        line_item_cell.font = Font(italic=True)
                        if has_parent:
                            # Indent child aggregators
                            line_item_cell.alignment = Alignment(indent=1)
                        else:
                            # No indentation for top-level aggregators
                            line_item_cell.alignment = Alignment(indent=0)
                    elif has_parent:
                        # Regular font with indentation for child line items
                        line_item_cell.font = Font()
                        line_item_cell.alignment = Alignment(indent=1)
                    else:
                        # Regular font for main line items
                        line_item_cell.font = Font()
                        line_item_cell.alignment = Alignment(indent=0)
                
                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = openpyxl.utils.get_column_letter(column[0].column)
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                    ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save the formatted workbook
        progress("    • Saving formatted Excel file...")
        wb.save(filepath)
        progress(f"    ✓ Excel file formatted and saved: {filepath}")
        return filepath

    def export_to_excel(self, financial_model: Dict[str, pd.DataFrame], 
                       sensitivity_model: Dict[str, pd.DataFrame], 
                       ticker: str, filename: str = None, schmoove_mode: bool = False, progress_callback=None) -> str:
        """
        Export financial models to Excel file with professional formatting for annual and quarterly sheets.
        If schmoove_mode is True, use parallel processing for formatting and allow higher resource usage.
        """
        def progress(msg):
            if progress_callback:
                progress_callback(msg)
            else:
                print(msg)
        
        import pandas as pd
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        import os
        from datetime import datetime
        
        progress("    • Initializing Excel export...")
        
        if schmoove_mode:
            try:
                import joblib
                import psutil
                # Set pandas and openpyxl to use more threads if possible
                pd.set_option('compute.use_numexpr', True)
                # Limit joblib to 75% of CPUs
                n_jobs = max(1, int(psutil.cpu_count(logical=True) * 0.75))
                # Limit memory usage to half of available RAM
                mem = psutil.virtual_memory()
                max_mem = int(mem.total * 0.5)
                progress(f"    • Schmoove mode enabled: Using {n_jobs} CPU cores")
            except ImportError:
                n_jobs = 1
                max_mem = None
        else:
            n_jobs = 1
            max_mem = None
        
        # Parent-child mapping for indentation following standard three-statement modeling
        parent_map = self._create_user_friendly_parent_map()
        
        # Ensure the Storage directory exists
        storage_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))), 'Storage')
        os.makedirs(storage_dir, exist_ok=True)
        
        if filename is None:
            filename = f"{ticker}_financial_model_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(storage_dir, filename)
        
        progress(f"    • Creating Excel file: {filename}")

        # Write all DataFrames to Excel using pandas (fast)
        progress("    • Writing financial data to Excel sheets...")
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            # Annual sheet
            annual_income = financial_model.get('annual_income_statement', pd.DataFrame())
            annual_balance = financial_model.get('annual_balance_sheet', pd.DataFrame())
            annual_cash_flow = financial_model.get('annual_cash_flow', pd.DataFrame())
            if not annual_income.empty or not annual_balance.empty or not annual_cash_flow.empty:
                progress("    • Writing annual financial statements...")
                stacked_annual = self._create_vertically_stacked_statement(annual_income, annual_balance, annual_cash_flow, period_type="Annual")
                # Remove empty columns before writing to Excel
                stacked_annual = self._remove_empty_columns(stacked_annual)
                # Remove formatting columns before writing to Excel
                stacked_annual = self._remove_formatting_columns(stacked_annual)
                stacked_annual.to_excel(writer, sheet_name='Annual Financial Statements', index=False)
            # Quarterly sheet
            quarterly_income = financial_model.get('quarterly_income_statement', pd.DataFrame())
            quarterly_balance = financial_model.get('quarterly_balance_sheet', pd.DataFrame())
            quarterly_cash_flow = financial_model.get('quarterly_cash_flow', pd.DataFrame())
            if not quarterly_income.empty or not quarterly_balance.empty or not quarterly_cash_flow.empty:
                progress("    • Writing quarterly financial statements...")
                stacked_quarterly = self._create_vertically_stacked_statement(quarterly_income, quarterly_balance, quarterly_cash_flow, period_type="Quarterly")
                # Remove empty columns before writing to Excel
                stacked_quarterly = self._remove_empty_columns(stacked_quarterly)
                # Remove formatting columns before writing to Excel
                stacked_quarterly = self._remove_formatting_columns(stacked_quarterly)
                stacked_quarterly.to_excel(writer, sheet_name='Quarterly Financial Statements', index=False)
            # Sensitivity and summary sheets
            progress("    • Writing sensitivity analysis and summary sheets...")
            for sheet_name, df in sensitivity_model.items():
                if not df.empty:
                    # Remove empty columns before writing to Excel
                    cleaned_df = self._remove_empty_columns(df)
                    cleaned_df.to_excel(writer, sheet_name=sheet_name.replace('_', ' ').title())
            # Summary sheet
            summary_data = {
                'Metric': ['Ticker', 'Model Created', 'Data Points', 'Scenarios Analyzed'],
                'Value': [
                    ticker,
                    datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    sum(len(df) for df in financial_model.values() if not df.empty),
                    len(sensitivity_model.get('case_summary', pd.DataFrame()))
                ]
            }
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='Summary', index=False)

        # Now open with openpyxl and apply formatting
        progress("    • Applying professional formatting...")
        wb = openpyxl.load_workbook(filepath)
        
        # Format annual and quarterly sheets
        for sheet in ['Annual Financial Statements', 'Quarterly Financial Statements']:
            if sheet in wb.sheetnames:
                ws = wb[sheet]
                
                # Find the key columns
                line_item_col = None
                
                for idx, cell in enumerate(ws[1]):
                    if cell.value == 'Line Item':
                        line_item_col = idx + 1
                        break
                
                if line_item_col is None:
                    continue
                
                # Apply formatting to each row based on line item content
                for row_num in range(2, ws.max_row + 1):
                    line_item_cell = ws.cell(row=row_num, column=line_item_col)
                    line_item_value = line_item_cell.value
                    
                    if line_item_value is None:
                        continue
                    
                    # Apply formatting based on line item content
                    if line_item_value.upper() in ['INCOME STATEMENT', 'BALANCE SHEET', 'CASH FLOW STATEMENT']:
                        # Bold for section headings
                        line_item_cell.font = Font(bold=True, size=12)
                        line_item_cell.alignment = Alignment(horizontal='left')
                    elif any(keyword in line_item_value for keyword in ['Total', 'Net', 'Gross', 'Income', 'Earnings', 'Cash at End']):
                        # Italic for aggregator line items
                        line_item_cell.font = Font(italic=True)
                        line_item_cell.alignment = Alignment(indent=0)
                    else:
                        # Regular font for main line items
                        line_item_cell.font = Font()
                        line_item_cell.alignment = Alignment(indent=0)
                
                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = openpyxl.utils.get_column_letter(column[0].column)
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                    ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save the formatted workbook
        progress("    • Saving formatted Excel file...")
        wb.save(filepath)
        progress(f"    ✓ Excel file saved: {filepath}")
        return filepath

    def _create_vertically_stacked_statement(self, income_df, balance_df, cash_flow_df, period_type):
        """
        Create a vertically stacked financial statement with proper formatting following standard three-statement modeling principles.
        Returns a clean DataFrame with line items and values, formatting will be applied directly in Excel.
        Each unique line item appears only once per sheet, with different reporting periods as columns.
        """
        import pandas as pd
        
        # Define the order and structure for each statement with proper parent-child relationships
        income_order = [
            # Income Statement Section Heading
            ('Income Statement', None),  # Section heading
            
            # Revenues Subheading
            ('Revenues', None),  # Subheading title
            ('Revenue Stream One', 'Revenues'),  # Child of Revenues
            ('Revenue Stream Two', 'Revenues'),  # Child of Revenues
            ('Total Revenues', 'Revenues'),  # Subtotal of Revenues
            
            # OpEx Subheading
            ('OpEx', None),  # Subheading title
            ('Cost of Goods Sold', 'OpEx'),  # Child of OpEx
            ('SG&A', 'OpEx'),  # Child of OpEx
            ('Misc', 'OpEx'),  # Child of OpEx
            ('Total OpEx', 'OpEx'),  # Subtotal of OpEx
            
            # EBITDA (Total Revenues - Total OpEx)
            ('EBITDA', None),  # Total Revenues - Total OpEx
            
            # Other expenses
            ('Depreciation and Amortization', None),
            ('Interest Expense', None),
            ('Tax Expense', None),
            
            # Net Income (EBITDA - Depreciation and Amortization - Interest Expense - Tax Expense)
            ('Net Income', None),  # EBITDA - Depreciation and Amortization - Interest Expense - Tax Expense
        ]
        balance_order = [
            ('CashAndCashEquivalents', 'Current Assets'),
            ('ShortTermInvestments', 'Current Assets'),
            ('AccountsReceivable', 'Current Assets'),
            ('Inventory', 'Current Assets'),
            ('PrepaidExpenses', 'Current Assets'),
            ('OtherCurrentAssets', 'Current Assets'),
            ('TotalCurrentAssets', None),
            ('PropertyPlantAndEquipmentNet', 'Non-Current Assets'),
            ('Goodwill', 'Non-Current Assets'),
            ('IntangibleAssetsNet', 'Non-Current Assets'),
            ('LongTermInvestments', 'Non-Current Assets'),
            ('DeferredTaxAssets', 'Non-Current Assets'),
            ('OtherLongTermAssets', 'Non-Current Assets'),
            ('TotalNonCurrentAssets', None),
            ('TotalAssets', None),
            ('AccountsPayable', 'Current Liabilities'),
            ('AccruedExpenses', 'Current Liabilities'),
            ('DeferredRevenue', 'Current Liabilities'),
            ('ShortTermDebt', 'Current Liabilities'),
            ('OtherCurrentLiabilities', 'Current Liabilities'),
            ('TotalCurrentLiabilities', None),
            ('LongTermDebt', 'Non-Current Liabilities'),
            ('DeferredTaxLiabilities', 'Non-Current Liabilities'),
            ('OtherLongTermLiabilities', 'Non-Current Liabilities'),
            ('TotalNonCurrentLiabilities', None),
            ('TotalLiabilities', None),
            ('CommonStock', 'Stockholders\' Equity'),
            ('AdditionalPaidInCapital', 'Stockholders\' Equity'),
            ('RetainedEarnings', 'Stockholders\' Equity'),
            ('AccumulatedOtherComprehensiveIncome', 'Stockholders\' Equity'),
            ('TreasuryStock', 'Stockholders\' Equity'),
            ('TotalStockholdersEquity', None),
            ('WorkingCapital', None),
            ('TotalDebt', None),
        ]
        cash_flow_order = [
            ('NetIncome', None),  # This will be consolidated with Income Statement Net Income
            ('DepreciationAndAmortization', 'Operating Adjustments'),
            ('StockBasedCompensation', 'Operating Adjustments'),
            ('DeferredIncomeTaxes', 'Operating Adjustments'),
            ('OperatingAdjustments', None),
            ('ChangeInAccountsReceivable', 'Working Capital Changes'),
            ('ChangeInInventory', 'Working Capital Changes'),
            ('ChangeInAccountsPayable', 'Working Capital Changes'),
            ('ChangeInDeferredRevenue', 'Working Capital Changes'),
            ('ChangeInOtherWorkingCapital', 'Working Capital Changes'),
            ('WorkingCapitalChanges', None),
            ('OtherOperatingActivities', 'Operating Activities'),
            ('OperatingActivities', None),
            ('NetCashFromOperatingActivities', None),
            ('CapitalExpenditures', 'Investing Activities'),
            ('Acquisitions', 'Investing Activities'),
            ('Investments', 'Investing Activities'),
            ('ProceedsFromInvestments', 'Investing Activities'),
            ('OtherInvestingActivities', 'Investing Activities'),
            ('InvestingActivities', None),
            ('NetCashFromInvestingActivities', None),
            ('ProceedsFromDebt', 'Financing Activities'),
            ('RepaymentsOfDebt', 'Financing Activities'),
            ('DividendsPaid', 'Financing Activities'),
            ('StockRepurchases', 'Financing Activities'),
            ('ProceedsFromStockIssuance', 'Financing Activities'),
            ('OtherFinancingActivities', 'Financing Activities'),
            ('FinancingActivities', None),
            ('NetCashFromFinancingActivities', None),
            ('EffectOfExchangeRateChanges', None),
            ('NetChangeInCash', None),
            ('CashAtBeginningOfPeriod', None),
            ('CashAtEndOfPeriod', None),
        ]
        
        # Track processed line items to avoid duplicates
        processed_line_items = set()
        
        # Helper to build section with hierarchical structure
        def build_section(section_title, order, df):
            rows = []
            
            # Insert section heading
            section_row = {
                'Line Item': section_title,
                'statement_type': section_title.split()[0].upper(),  # INCOME, BALANCE, CASH
                'is_section_heading': True,
                'parent': None,
                'is_aggregate': False
            }
            
            # Add all available dates as columns to the section heading
            if not df.empty:
                for date in df.index:
                    section_row[date] = ''  # Empty values for section headings
            
            rows.append(section_row)
            
            # Create a mapping of parent concepts to their children
            parent_children_map = {}
            for item, parent in order:
                if parent is not None:
                    if parent not in parent_children_map:
                        parent_children_map[parent] = []
                    parent_children_map[parent].append(item)
            
            # Process each line item in hierarchical order
            for item, parent in order:
                if item in df.columns:
                    # Convert technical concept name to user-friendly title
                    friendly_title = self._get_user_friendly_title(item)
                    
                    # Handle duplicate line items that appear in multiple statements
                    # For Net Income, we want to keep it in the Income Statement and skip it in Cash Flow
                    if friendly_title == 'Net income' and section_title == 'CASH FLOW STATEMENT':
                        # Skip Net Income in Cash Flow Statement since it's already in Income Statement
                        continue
                    
                    # Check if this line item has already been processed
                    if friendly_title in processed_line_items:
                        # Skip duplicates
                        continue
                    
                    # Determine if this is an aggregate/total line item
                    is_aggregate = (parent is None and 
                                  (item.lower().startswith('total') or 
                                   'Net' in item or 'Gross' in item or 
                                   'Income' in item or 'Earnings' in item or 
                                   'OperatingIncome' in item or 
                                   'CashAtEndOfPeriod' == item))
                    
                    # If this item has children, create a parent row first
                    if item in parent_children_map:
                        # Create parent row
                        parent_friendly_title = self._get_user_friendly_title(item)
                        parent_row = {
                            'Line Item': parent_friendly_title,
                            'statement_type': section_title.split()[0].upper(),
                            'is_section_heading': False,
                            'parent': parent,
                            'is_aggregate': True  # Parent rows are aggregators
                        }
                        
                        # Add values for each period as columns
                        for date in df.index:
                            parent_row[date] = df.loc[date, item]
                        
                        rows.append(parent_row)
                        processed_line_items.add(parent_friendly_title)
                        
                        # Now add all children under this parent
                        for child_item in parent_children_map[item]:
                            if child_item in df.columns:
                                child_friendly_title = self._get_user_friendly_title(child_item)
                                
                                # Skip if already processed
                                if child_friendly_title in processed_line_items:
                                    continue
                                
                                # Create child row
                                child_row = {
                                    'Line Item': child_friendly_title,
                                    'statement_type': section_title.split()[0].upper(),
                                    'is_section_heading': False,
                                    'parent': item,  # Reference to parent
                                    'is_aggregate': False  # Children are not aggregators
                                }
                                
                                # Add values for each period as columns
                                for date in df.index:
                                    child_row[date] = df.loc[date, child_item]
                                
                                rows.append(child_row)
                                processed_line_items.add(child_friendly_title)
                    
                    # If this item doesn't have children and hasn't been processed as a child
                    elif friendly_title not in processed_line_items:
                        # Create a single row for this line item
                        row = {
                            'Line Item': friendly_title,
                            'statement_type': section_title.split()[0].upper(),
                            'is_section_heading': False,
                            'parent': parent,
                            'is_aggregate': is_aggregate
                        }
                        
                        # Add values for each period as columns
                        for date in df.index:
                            row[date] = df.loc[date, item]
                        
                        rows.append(row)
                        processed_line_items.add(friendly_title)
            
            return rows
        
        # Build all sections
        rows = []
        rows += build_section('INCOME STATEMENT', income_order, income_df)
        
        # Add blank row between sections
        if rows:  # Only add blank row if we have data
            blank_row = {'Line Item': '', 'statement_type': '', 'is_section_heading': False, 'parent': None, 'is_aggregate': False}
            # Add empty values for all available dates
            if not income_df.empty:
                for date in income_df.index:
                    blank_row[date] = ''
            rows.append(blank_row)
        
        rows += build_section('BALANCE SHEET', balance_order, balance_df)
        
        # Add blank row between sections
        if rows:  # Only add blank row if we have data
            blank_row = {'Line Item': '', 'statement_type': '', 'is_section_heading': False, 'parent': None, 'is_aggregate': False}
            # Add empty values for all available dates
            if not balance_df.empty:
                for date in balance_df.index:
                    blank_row[date] = ''
            rows.append(blank_row)
        
        rows += build_section('CASH FLOW STATEMENT', cash_flow_order, cash_flow_df)
        
        # Convert to DataFrame
        if rows:
            df = pd.DataFrame(rows)
            df = df.fillna('')
            # Remove empty columns before returning
            df = self._remove_empty_columns(df)
            return df
        else:
            return pd.DataFrame()

    def _extract_xbrl_from_filing(self, filing_row, cik, progress_callback=None):
        """
        Extract XBRL data from a specific filing.
        
        Args:
            filing_row: DataFrame row containing filing information
            cik: Company CIK number
            progress_callback: Optional callback function for progress updates
            
        Returns:
            Dict: Extracted XBRL facts
        """
        def progress(msg):
            if progress_callback:
                progress_callback(msg)
            else:
                print(msg)
        
        try:
            accession_number = filing_row['accessionNumber']
            cik_dir = str(int(cik))
            accession_clean = accession_number.replace('-', '')
            filing_dir_url = f"https://www.sec.gov/Archives/edgar/data/{cik_dir}/{accession_clean}/"
            
            # List files in the directory
            progress("        • Accessing filing directory...")
            dir_response = self.session.get(filing_dir_url)
            if dir_response.status_code != 200:
                progress(f"        ✗ Could not access filing directory: {filing_dir_url}")
                return None
            
            # Find .xml files (XBRL instance docs)
            import re
            xml_files = re.findall(r'href="([^"]+\.xml)"', dir_response.text)
            xbrl_instance_file = None
            
            progress(f"        • Found {len(xml_files)} XML files")
            
            # Prefer files ending with _htm.xml and not FilingSummary.xml
            for xml_file in xml_files:
                if xml_file.endswith('_htm.xml') and 'FilingSummary' not in xml_file:
                    xbrl_instance_file = xml_file
                    break
            
            if xbrl_instance_file is None:
                # Fallback: pick the first .xml file that's not FilingSummary
                for xml_file in xml_files:
                    if 'FilingSummary' not in xml_file:
                        xbrl_instance_file = xml_file
                        break
            
            if not xbrl_instance_file:
                progress(f"        ✗ No XBRL instance document found")
                return None
            
            # Construct URL
            if xbrl_instance_file.startswith('/'):
                xbrl_url = f"https://www.sec.gov{xbrl_instance_file}"
            else:
                xbrl_url = f"https://www.sec.gov/Archives/edgar/data/{cik_dir}/{accession_clean}/{xbrl_instance_file}"
            
            progress(f"        • Parsing XBRL: {xbrl_instance_file}")
            
            # Use Arelle to parse the XBRL instance document
            from arelle import Cntlr
            cntlr = Cntlr.Cntlr(logFileName=None)
            model_xbrl = cntlr.modelManager.load(xbrl_url)
            
            # Extract facts from the XBRL model
            facts_from_xml = {}
            fact_count = 0
            
            progress(f"        • Extracting facts from XBRL model...")
            for fact in model_xbrl.facts:
                try:
                    concept_name = str(fact.qname)
                    value = fact.value
                    context = fact.context
                    
                    if context is not None:
                        period = context.period
                        if period is not None:
                            end_date = getattr(period, 'endDate', None)
                            if end_date is None:
                                end_date = getattr(period, 'end', None)
                            if end_date is None:
                                end_date = getattr(context, 'endDate', None)
                            
                            if end_date:
                                if concept_name not in facts_from_xml:
                                    facts_from_xml[concept_name] = []
                                facts_from_xml[concept_name].append({
                                    'value': value,
                                    'end_date': end_date,
                                    'context': context,
                                    'filing_type': filing_row['form'],
                                    'filing_date': filing_row['filingDate']
                                })
                                fact_count += 1
                except Exception as e:
                    continue
            
            progress(f"        ✓ Extracted {len(facts_from_xml)} concepts ({fact_count} total facts)")
            return facts_from_xml
            
        except Exception as e:
            progress(f"        ✗ Error extracting XBRL from filing: {str(e)}")
            return None

    def _run_discrepancy_checks(self, annual_data, quarterly_data):
        """
        Run discrepancy checks between annual and quarterly data for key metrics.
        
        Args:
            annual_data: Annual data from 10-K filings
            quarterly_data: Quarterly data from 10-Q filings
        """
        print("\n" + "="*60)
        print("DISCREPANCY CHECKS: Annual vs Quarterly Data")
        print("="*60)
        
        # Key metrics to check
        key_metrics = [
            'us-gaap:NetIncomeLoss',
            'us-gaap:RevenueFromContractWithCustomerExcludingAssessedTax',
            'us-gaap:GrossProfit',
            'us-gaap:OperatingIncomeLoss'
        ]
        
        for metric in key_metrics:
            if metric in annual_data and metric in quarterly_data:
                print(f"\nChecking: {metric}")
                
                # Get annual values
                annual_values = {}
                for item in annual_data[metric]:
                    date = item['end_date']
                    value = float(item['value'])
                    annual_values[date] = value
                
                # Get quarterly values and sum them by year
                quarterly_sums = {}
                for item in quarterly_data[metric]:
                    date = item['end_date']
                    year = date.year
                    value = float(item['value'])
                    
                    if year not in quarterly_sums:
                        quarterly_sums[year] = 0
                    quarterly_sums[year] += value
                
                # Compare annual vs quarterly sums
                for year, quarterly_sum in quarterly_sums.items():
                    # Find annual value for this year
                    annual_value = None
                    for date, value in annual_values.items():
                        if date.year == year:
                            annual_value = value
                            break
                    
                    if annual_value is not None:
                        difference = abs(annual_value - quarterly_sum)
                        difference_pct = (difference / annual_value * 100) if annual_value != 0 else 0
                        
                        print(f"  {year}: Annual={annual_value:,.0f}, Quarterly Sum={quarterly_sum:,.0f}")
                        print(f"    Difference: {difference:,.0f} ({difference_pct:.2f}%)")
                        
                        if difference_pct > 5:  # Flag differences > 5%
                            print(f"    SIGNIFICANT DISCREPANCY DETECTED!")
                        elif difference_pct > 1:  # Flag differences > 1%
                            print(f"    Minor discrepancy detected")
                        else:
                            print(f"    Data consistent")
            else:
                print(f"\nSkipping {metric}: Not available in both annual and quarterly data")

    def _create_model_from_comprehensive_data(self, comprehensive_facts, annual_data, quarterly_data, enhanced_fuzzy_matching=True, progress_callback=None):
        """
        Create financial model from comprehensive 10-K and 10-Q data.
        
        Args:
            comprehensive_facts: Combined facts from all filings (used for fuzzy matching)
            annual_data: Annual data from 10-K filings
            quarterly_data: Quarterly data from 10-Q filings
            enhanced_fuzzy_matching: Whether to include non-GAAP to GAAP mapping
            progress_callback: Optional callback function for progress updates
            
        Returns:
            Dict: Financial model with annual and quarterly data
        """
        # Define progress function for this method
        def progress(msg):
            if progress_callback:
                progress_callback(msg)
            else:
                print(msg)
        
        financial_model = {
            'annual_income_statement': pd.DataFrame(),
            'annual_balance_sheet': pd.DataFrame(),
            'annual_cash_flow': pd.DataFrame(),
            'quarterly_income_statement': pd.DataFrame(),
            'quarterly_balance_sheet': pd.DataFrame(),
            'quarterly_cash_flow': pd.DataFrame()
        }
        
        # Get all available concepts from comprehensive_facts for fuzzy matching
        available_concepts = list(comprehensive_facts.keys())
        print(f"\nUsing fuzzy matching to find {len(available_concepts)} available concepts...")
        if enhanced_fuzzy_matching:
            print("Enhanced fuzzy matching enabled: Will include non-GAAP to GAAP mapping")
            progress("    • Enhanced fuzzy matching enabled (includes non-GAAP to GAAP mapping)")
        else:
            print("Standard fuzzy matching: GAAP concepts only")
            progress("    • Standard fuzzy matching enabled (GAAP concepts only)")
        
        progress("    • Applying standard three-statement financial modeling principles...")
        
        # Define desired metrics with multiple possible tags for each concept
        # Following standard three-statement financial modeling principles
        
        # CORE GAAP CONCEPTS - Essential line items that should always be included
        # These are the fundamental building blocks of financial statements
        
        # INCOME STATEMENT - Core GAAP concepts (top priority)
        core_income_statement_metrics = {
            # Revenue (top line)
            'Revenue': ['us-gaap:Revenues', 'us-gaap:RevenueFromContractWithCustomerExcludingAssessedTax', 'us-gaap:SalesRevenueNet'],
            
            # Cost of goods sold
            'CostOfGoodsSold': ['us-gaap:CostOfRevenue', 'us-gaap:CostOfGoodsAndServicesSold'],
            
            # Gross profit
            'GrossProfit': ['us-gaap:GrossProfit', 'us-gaap:GrossProfitLoss'],
            
            # Core operating expenses
            'ResearchAndDevelopmentExpense': ['us-gaap:ResearchAndDevelopmentExpense'],
            'SellingGeneralAndAdministrativeExpense': ['us-gaap:SellingGeneralAndAdministrativeExpense'],
            
            # Operating income
            'OperatingIncome': ['us-gaap:OperatingIncomeLoss'],
            
            # Interest and taxes
            'InterestExpense': ['us-gaap:InterestExpense'],
            'IncomeTaxExpense': ['us-gaap:IncomeTaxExpenseBenefit'],
            
            # Net income (bottom line)
            'NetIncome': ['us-gaap:NetIncomeLoss'],
            
            # Earnings per share
            'EarningsPerShareBasic': ['us-gaap:EarningsPerShareBasic'],
            'EarningsPerShareDiluted': ['us-gaap:EarningsPerShareDiluted']
        }
        
        # BALANCE SHEET - Core GAAP concepts
        core_balance_sheet_metrics = {
            # Core current assets
            'CashAndCashEquivalents': ['us-gaap:CashAndCashEquivalentsAtCarryingValue'],
            'AccountsReceivable': ['us-gaap:AccountsReceivableNetCurrent'],
            'Inventory': ['us-gaap:InventoryNet'],
            'TotalCurrentAssets': ['us-gaap:AssetsCurrent'],
            
            # Core non-current assets
            'PropertyPlantAndEquipmentNet': ['us-gaap:PropertyPlantAndEquipmentNet'],
            'Goodwill': ['us-gaap:Goodwill'],
            'TotalAssets': ['us-gaap:Assets'],
            
            # Core current liabilities
            'AccountsPayable': ['us-gaap:AccountsPayableCurrent'],
            'TotalCurrentLiabilities': ['us-gaap:LiabilitiesCurrent'],
            
            # Core non-current liabilities
            'LongTermDebt': ['us-gaap:LongTermDebtNoncurrent'],
            'TotalLiabilities': ['us-gaap:Liabilities'],
            
            # Core equity
            'RetainedEarnings': ['us-gaap:RetainedEarningsAccumulatedDeficit'],
            'TotalStockholdersEquity': ['us-gaap:StockholdersEquity']
        }
        
        # CASH FLOW STATEMENT - Core GAAP concepts
        core_cash_flow_metrics = {
            # Operating activities
            'NetIncome': ['us-gaap:NetIncomeLoss'],
            'DepreciationAndAmortization': ['us-gaap:DepreciationAndAmortization'],
            'NetCashFromOperatingActivities': ['us-gaap:NetCashProvidedByUsedInOperatingActivities'],
            
            # Investing activities
            'CapitalExpenditures': ['us-gaap:PaymentsToAcquirePropertyPlantAndEquipment'],
            'NetCashFromInvestingActivities': ['us-gaap:NetCashProvidedByUsedInInvestingActivities'],
            
            # Financing activities
            'DividendsPaid': ['us-gaap:PaymentsOfDividends'],
            'NetCashFromFinancingActivities': ['us-gaap:NetCashProvidedByUsedInFinancingActivities'],
            
            # Net change
            'NetChangeInCash': ['us-gaap:CashCashEquivalentsRestrictedCashAndRestrictedCashEquivalentsPeriodIncreaseDecreaseIncludingExchangeRateEffect']
        }
        
        # ADDITIONAL GAAP CONCEPTS - Up to 20 additional items that are GAAP-compliant and properly categorized
        additional_income_statement_metrics = {
            # Additional operating items
            'DepreciationAndAmortization': ['us-gaap:DepreciationAndAmortization'],
            'StockBasedCompensationExpense': ['us-gaap:StockBasedCompensationExpense'],
            'RestructuringCharges': ['us-gaap:RestructuringCharges'],
            'ImpairmentCharges': ['us-gaap:ImpairmentCharges'],
            
            # Additional non-operating items
            'InterestIncome': ['us-gaap:InterestIncome'],
            'GainLossOnSaleOfAssets': ['us-gaap:GainLossOnSaleOfAssets'],
            'ForeignCurrencyGainLoss': ['us-gaap:ForeignCurrencyGainLoss'],
            'OtherIncomeExpense': ['us-gaap:OtherIncomeExpenseNet'],
            
            # Additional metrics
            'IncomeBeforeTaxes': ['us-gaap:IncomeLossFromContinuingOperationsBeforeIncomeTaxes'],
            'WeightedAverageSharesBasic': ['us-gaap:WeightedAverageNumberOfSharesOutstandingBasic'],
            'WeightedAverageSharesDiluted': ['us-gaap:WeightedAverageNumberOfSharesOutstandingDiluted']
        }
        
        additional_balance_sheet_metrics = {
            # Additional current assets
            'ShortTermInvestments': ['us-gaap:ShortTermInvestments'],
            'PrepaidExpenses': ['us-gaap:PrepaidExpenseAndOtherAssetsCurrent'],
            'OtherCurrentAssets': ['us-gaap:OtherAssetsCurrent'],
            
            # Additional non-current assets
            'IntangibleAssetsNet': ['us-gaap:IntangibleAssetsNetExcludingGoodwill'],
            'LongTermInvestments': ['us-gaap:InvestmentsNoncurrent'],
            'DeferredTaxAssets': ['us-gaap:DeferredTaxAssetsNet'],
            'OtherLongTermAssets': ['us-gaap:OtherAssetsNoncurrent'],
            'TotalNonCurrentAssets': ['us-gaap:AssetsNoncurrent'],
            
            # Additional current liabilities
            'AccruedExpenses': ['us-gaap:AccruedLiabilitiesCurrent'],
            'DeferredRevenue': ['us-gaap:ContractWithCustomerLiabilityCurrent'],
            'ShortTermDebt': ['us-gaap:ShortTermBorrowings'],
            'OtherCurrentLiabilities': ['us-gaap:OtherLiabilitiesCurrent'],
            
            # Additional non-current liabilities
            'DeferredTaxLiabilities': ['us-gaap:DeferredTaxLiabilitiesNet'],
            'OtherLongTermLiabilities': ['us-gaap:OtherLiabilitiesNoncurrent'],
            'TotalNonCurrentLiabilities': ['us-gaap:LiabilitiesNoncurrent'],
            
            # Additional equity
            'CommonStock': ['us-gaap:CommonStockValue'],
            'AdditionalPaidInCapital': ['us-gaap:AdditionalPaidInCapital'],
            'AccumulatedOtherComprehensiveIncome': ['us-gaap:AccumulatedOtherComprehensiveIncomeLossNetOfTax'],
            'TreasuryStock': ['us-gaap:TreasuryStockValue'],
            
            # Additional calculated metrics
            'WorkingCapital': ['us-gaap:WorkingCapital']
        }
        
        additional_cash_flow_metrics = {
            # Additional operating activities
            'StockBasedCompensation': ['us-gaap:StockBasedCompensationExpense'],
            'DeferredIncomeTaxes': ['us-gaap:DeferredIncomeTaxExpenseBenefit'],
            'ChangeInAccountsReceivable': ['us-gaap:IncreaseDecreaseInAccountsReceivable'],
            'ChangeInInventory': ['us-gaap:IncreaseDecreaseInInventories'],
            'ChangeInAccountsPayable': ['us-gaap:IncreaseDecreaseInAccountsPayable'],
            'ChangeInDeferredRevenue': ['us-gaap:IncreaseDecreaseInContractWithCustomerLiability'],
            'OtherOperatingActivities': ['us-gaap:OtherOperatingActivitiesCashFlowStatement'],
            
            # Additional investing activities
            'Acquisitions': ['us-gaap:PaymentsToAcquireBusinessesNetOfCashAcquired'],
            'Investments': ['us-gaap:PaymentsToAcquireInvestments'],
            'ProceedsFromInvestments': ['us-gaap:ProceedsFromSaleMaturityAndCollectionsOfInvestments'],
            'OtherInvestingActivities': ['us-gaap:OtherInvestingActivitiesCashFlowStatement'],
            
            # Additional financing activities
            'ProceedsFromDebt': ['us-gaap:ProceedsFromIssuanceOfLongTermDebt'],
            'RepaymentsOfDebt': ['us-gaap:RepaymentsOfLongTermDebt'],
            'StockRepurchases': ['us-gaap:PaymentsForRepurchaseOfCommonStock'],
            'ProceedsFromStockIssuance': ['us-gaap:ProceedsFromIssuanceOfCommonStock'],
            'OtherFinancingActivities': ['us-gaap:OtherFinancingActivitiesCashFlowStatement'],
            
            # Additional metrics
            'EffectOfExchangeRateChanges': ['us-gaap:EffectOfExchangeRateOnCashCashEquivalentsRestrictedCashAndRestrictedCashEquivalents'],
            'CashAtBeginningOfPeriod': ['us-gaap:CashCashEquivalentsRestrictedCashAndRestrictedCashEquivalentsPeriodBeginningBalance'],
            'CashAtEndOfPeriod': ['us-gaap:CashCashEquivalentsRestrictedCashAndRestrictedCashEquivalentsPeriodEndingBalance']
        }
        
        # Function to selectively add additional metrics (up to 20 total additional items)
        def add_selective_additional_metrics(core_mapping, additional_metrics, available_concepts, enhanced_fuzzy_matching):
            """Add up to 20 additional GAAP-compliant metrics that are available in the data."""
            additional_mapping = {}
            total_additional = 0
            max_additional = 20
            
            progress(f"    • Adding up to {max_additional} additional GAAP-compliant line items...")
            
            for concept_name, possible_tags in additional_metrics.items():
                if total_additional >= max_additional:
                    break
                    
                for tag in possible_tags:
                    # Check if this concept is available and not already in core mapping
                    if tag in available_concepts and concept_name not in core_mapping:
                        mapping = self._fuzzy_match_concepts({concept_name: tag}, available_concepts, similarity_threshold=0.85, enhanced_fuzzy_matching=enhanced_fuzzy_matching)
                        if mapping:
                            additional_mapping.update(mapping)
                            total_additional += 1
                            progress(f"      ✓ Added: {concept_name}")
                            break
            
            progress(f"    • Added {total_additional} additional line items")
            return additional_mapping
        
        # Start with core metrics
        progress("    • Matching core GAAP concepts...")
        
        # Match core income statement concepts
        income_mapping = {}
        for concept_name, possible_tags in core_income_statement_metrics.items():
            for tag in possible_tags:
                mapping = self._fuzzy_match_concepts({concept_name: tag}, available_concepts, similarity_threshold=0.85, enhanced_fuzzy_matching=enhanced_fuzzy_matching)
                if mapping:
                    income_mapping.update(mapping)
                    break
        
        # Match core balance sheet concepts
        balance_mapping = {}
        for concept_name, possible_tags in core_balance_sheet_metrics.items():
            for tag in possible_tags:
                mapping = self._fuzzy_match_concepts({concept_name: tag}, available_concepts, similarity_threshold=0.85, enhanced_fuzzy_matching=enhanced_fuzzy_matching)
                if mapping:
                    balance_mapping.update(mapping)
                    break
        
        # Match core cash flow concepts
        cash_flow_mapping = {}
        for concept_name, possible_tags in core_cash_flow_metrics.items():
            for tag in possible_tags:
                mapping = self._fuzzy_match_concepts({concept_name: tag}, available_concepts, similarity_threshold=0.85, enhanced_fuzzy_matching=enhanced_fuzzy_matching)
                if mapping:
                    cash_flow_mapping.update(mapping)
                    break
        
        # Add selective additional metrics
        income_mapping.update(add_selective_additional_metrics(income_mapping, additional_income_statement_metrics, available_concepts, enhanced_fuzzy_matching))
        balance_mapping.update(add_selective_additional_metrics(balance_mapping, additional_balance_sheet_metrics, available_concepts, enhanced_fuzzy_matching))
        cash_flow_mapping.update(add_selective_additional_metrics(cash_flow_mapping, additional_cash_flow_metrics, available_concepts, enhanced_fuzzy_matching))
        
        # Create annual dataframes (from annual_data - 10-K filings)
        print(f"\nCreating annual financial statements...")
        for statement, mapping in [('income_statement', income_mapping),
                                 ('balance_sheet', balance_mapping),
                                 ('cash_flow', cash_flow_mapping)]:
            
            df = self._create_dataframe_from_data(annual_data, mapping, 'annual')
            financial_model[f'annual_{statement}'] = df
            print(f"Annual {statement}: {len(df)} rows extracted")
        
        # Create quarterly dataframes (from quarterly_data - 10-Q filings)
        print(f"\nCreating quarterly financial statements...")
        for statement, mapping in [('income_statement', income_mapping),
                                 ('balance_sheet', balance_mapping),
                                 ('cash_flow', cash_flow_mapping)]:
            
            df = self._create_dataframe_from_data(quarterly_data, mapping, 'quarterly')
            financial_model[f'quarterly_{statement}'] = df
            print(f"Quarterly {statement}: {len(df)} rows extracted")
        
        return financial_model

    def _create_dataframe_from_data(self, data, mapping, period_type):
        """
        Create DataFrame from annual or quarterly data using concept mapping.
        Now includes comprehensive data validation to filter out non-financial data.
        
        Args:
            data: Annual or quarterly data dictionary
            mapping: Dictionary mapping concept names to actual XBRL tags
            period_type: 'annual' or 'quarterly'
            
        Returns:
            pd.DataFrame: DataFrame with validated financial data
        """
        return self._extract_and_validate_financial_data(data, mapping, period_type)

    def _extract_and_validate_financial_data(self, data, mapping, period_type):
        """
        Extract financial data with comprehensive validation to filter out non-financial data.
        
        Args:
            data: Annual or quarterly data dictionary
            mapping: Dictionary mapping concept names to actual XBRL tags
            period_type: 'annual' or 'quarterly'
            
        Returns:
            pd.DataFrame: DataFrame with validated financial data
        """
        df_data = {}
        mapped_tags = set(mapping.values())
        
        # Collect all data points first, then filter by quarters
        all_data_points = []
        validation_stats = {
            'total_points': 0,
            'valid_points': 0,
            'rejected_points': 0,
            'rejected_reasons': {}
        }
        
        # 1. Add mapped concepts with validation
        for concept_name, actual_tag in mapping.items():
            if actual_tag in data:
                for item in data[actual_tag]:
                    validation_stats['total_points'] += 1
                    date = item['end_date']
                    
                    try:
                        value = float(item['value'])
                        unit = item.get('unit', '')
                        
                        # Validate the data point
                        if self._validate_financial_data(concept_name, value, unit):
                            all_data_points.append({
                                'date': date,
                                'concept': concept_name,
                                'value': value
                            })
                            validation_stats['valid_points'] += 1
                        else:
                            validation_stats['rejected_points'] += 1
                            reason = f"Non-financial data: {concept_name} = {value}"
                            validation_stats['rejected_reasons'][reason] = validation_stats['rejected_reasons'].get(reason, 0) + 1
                            
                    except (ValueError, TypeError):
                        validation_stats['rejected_points'] += 1
                        reason = f"Invalid numeric value: {concept_name} = {item.get('value', 'N/A')}"
                        validation_stats['rejected_reasons'][reason] = validation_stats['rejected_reasons'].get(reason, 0) + 1
                        continue
        
        # 2. Add unmapped tags as their own columns (with validation)
        for tag in data:
            if tag not in mapped_tags:
                for item in data[tag]:
                    validation_stats['total_points'] += 1
                    date = item['end_date']
                    
                    try:
                        value = float(item['value'])
                        unit = item.get('unit', '')
                        
                        # Validate the data point
                        if self._validate_financial_data(tag, value, unit):
                            all_data_points.append({
                                'date': date,
                                'concept': tag,
                                'value': value
                            })
                            validation_stats['valid_points'] += 1
                        else:
                            validation_stats['rejected_points'] += 1
                            reason = f"Non-financial data: {tag} = {value}"
                            validation_stats['rejected_reasons'][reason] = validation_stats['rejected_reasons'].get(reason, 0) + 1
                            
                    except (ValueError, TypeError):
                        validation_stats['rejected_points'] += 1
                        reason = f"Invalid numeric value: {tag} = {item.get('value', 'N/A')}"
                        validation_stats['rejected_reasons'][reason] = validation_stats['rejected_reasons'].get(reason, 0) + 1
                        continue
        
        # Log validation statistics
        if hasattr(self, 'progress_callback') and self.progress_callback:
            self.progress_callback(f"    • Data validation: {validation_stats['valid_points']}/{validation_stats['total_points']} points passed validation")
            if validation_stats['rejected_points'] > 0:
                self.progress_callback(f"    • Rejected {validation_stats['rejected_points']} non-financial data points")
                # Log top rejection reasons
                top_reasons = sorted(validation_stats['rejected_reasons'].items(), key=lambda x: x[1], reverse=True)[:3]
                for reason, count in top_reasons:
                    self.progress_callback(f"      - {reason}: {count} occurrences")
        
        if all_data_points:
            # Merge line items with same name but different dates
            all_data_points = self._merge_line_items_by_date(all_data_points, self.progress_callback if hasattr(self, 'progress_callback') else None)
            
            # Sort by date (most recent first)
            all_data_points.sort(key=lambda x: x['date'], reverse=True)
            
            # Filter to requested number of periods
            if period_type == 'annual':
                # For annual data, limit to the number of years needed
                years_needed = max(1, (self.current_quarters + 3) // 4)
                # Get unique years and take the most recent ones
                unique_years = []
                for point in all_data_points:
                    year = point['date'].year
                    if year not in unique_years:
                        unique_years.append(year)
                        if len(unique_years) >= years_needed:
                            break
                
                # Filter data points to only include the selected years
                filtered_points = [point for point in all_data_points if point['date'].year in unique_years]
            else:
                # For quarterly data, limit to the requested number of quarters
                # Get unique quarters and take the most recent ones
                unique_quarters = []
                for point in all_data_points:
                    quarter_key = (point['date'].year, point['date'].month)
                    if quarter_key not in unique_quarters:
                        unique_quarters.append(quarter_key)
                        if len(unique_quarters) >= self.current_quarters:
                            break
                
                # Filter data points to only include the selected quarters
                filtered_points = [point for point in all_data_points if (point['date'].year, point['date'].month) in unique_quarters]
            
            # Convert filtered data to DataFrame format
            for point in filtered_points:
                date_str = point['date'].strftime('%Y-%m-%d')
                if date_str not in df_data:
                    df_data[date_str] = {}
                df_data[date_str][point['concept']] = point['value']
        
        if df_data:
            df = pd.DataFrame.from_dict(df_data, orient='index')
            df.index = pd.to_datetime(df.index)
            df = df.sort_index()
            return df
        else:
            return pd.DataFrame()

    def _fuzzy_match_concepts(self, desired_concepts: Dict[str, str], available_concepts: List[str], 
                             similarity_threshold: float = 0.85, enhanced_fuzzy_matching: bool = True) -> Dict[str, str]:
        """
        Use fuzzy matching to find the best matches between desired XBRL concepts and available concepts.
        Also, for any non-US GAAP available concept, map it to the closest US GAAP tag if similarity > 0.7.
        Limited to 20 highest scoring mappings per financial statement section.
        Optimized: uses rapidfuzz if available, caches results, and parallelizes matching.
        
        Args:
            desired_concepts: Dictionary of concept names to desired XBRL tags
            available_concepts: List of available XBRL concepts
            similarity_threshold: Minimum similarity score for matching (default: 0.85)
            enhanced_fuzzy_matching: Whether to include non-GAAP to GAAP mapping (default: True)
            
        Returns:
            Dict[str, str]: Mapping of concept names to matched XBRL tags
        """
        concept_mapping = {}
        us_gaap_tags = [c for c in available_concepts if str(c).startswith('us-gaap:')]
        non_us_gaap_tags = [c for c in available_concepts if not str(c).startswith('us-gaap:')]
        cache_key = (tuple(sorted(desired_concepts.items())), tuple(sorted(available_concepts)), enhanced_fuzzy_matching)
        if cache_key in self._fuzzy_match_cache:
            return self._fuzzy_match_cache[cache_key].copy()
        def match_one(concept_name, desired_tag):
            best_match = None
            best_score = 0
            desired_main = desired_tag.split(':')[-1] if ':' in desired_tag else desired_tag
            if RAPIDFUZZ_AVAILABLE:
                # Use rapidfuzz for fast similarity
                matches = rf_process.extract(desired_main, [str(ac).split(':')[-1] for ac in available_concepts], scorer=fuzz.ratio, limit=3)
                for match_main, score, idx in matches:
                    score = score / 100.0
                    available_concept_str = str(available_concepts[idx])
                    if available_concept_str == desired_tag:
                        return (concept_name, available_concept_str, 1.0)
                    if score > best_score:
                        best_score = score
                        best_match = available_concept_str
            else:
                for available_concept in available_concepts:
                    available_concept_str = str(available_concept)
                    available_main = available_concept_str.split(':')[-1] if ':' in available_concept_str else available_concept_str
                    if available_concept_str == desired_tag:
                        return (concept_name, available_concept_str, 1.0)
                    if available_main == desired_main:
                        return (concept_name, available_concept_str, 0.95)
                    similarity = SequenceMatcher(None, available_main.lower(), desired_main.lower()).ratio()
                    if desired_main.lower() in available_main.lower() or available_main.lower() in desired_main.lower():
                        similarity = max(similarity, 0.8)
                    if similarity > best_score:
                        best_score = similarity
                        best_match = available_concept_str
            return (concept_name, best_match, best_score)
        # Parallelize matching when joblib is available, else sequential fallback
        if JOBLIB_AVAILABLE:
            results = Parallel(n_jobs=-1)(delayed(match_one)(concept_name, desired_tag) for concept_name, desired_tag in desired_concepts.items())
        else:
            results = [match_one(concept_name, desired_tag) for concept_name, desired_tag in desired_concepts.items()]
        for concept_name, best_match, best_score in results:
            if best_score >= similarity_threshold and best_match:
                concept_mapping[concept_name] = best_match
        
        # Only include non-GAAP to GAAP mapping if enhanced_fuzzy_matching is enabled
        if enhanced_fuzzy_matching:
            # --- New: For each non-US GAAP tag, map to closest US GAAP tag if similarity > 0.7 ---
            def match_non_gaap(non_gaap):
                non_gaap_str = str(non_gaap)
                non_gaap_main = non_gaap_str.split(':')[-1] if ':' in non_gaap_str else non_gaap_str
                best_usgaap = None
                best_score = 0
                if RAPIDFUZZ_AVAILABLE:
                    matches = rf_process.extract(non_gaap_main, [str(ug).split(':')[-1] for ug in us_gaap_tags], scorer=fuzz.ratio, limit=3)
                    for match_main, score, idx in matches:
                        score = score / 100.0
                        usgaap_str = str(us_gaap_tags[idx])
                        if score > best_score:
                            best_score = score
                            best_usgaap = usgaap_str
                else:
                    for usgaap in us_gaap_tags:
                        usgaap_str = str(usgaap)
                        usgaap_main = usgaap_str.split(':')[-1] if ':' in usgaap_str else usgaap_str
                        similarity = SequenceMatcher(None, non_gaap_main.lower(), usgaap_main.lower()).ratio()
                        if non_gaap_main.lower() in usgaap_main.lower() or usgaap_main.lower() in non_gaap_main.lower():
                            similarity = max(similarity, 0.8)
                        if similarity > best_score:
                            best_score = similarity
                            best_usgaap = usgaap_str
                return (non_gaap_str, best_usgaap, best_score)
            if JOBLIB_AVAILABLE:
                potential_mappings = Parallel(n_jobs=-1)(delayed(match_non_gaap)(non_gaap) for non_gaap in non_us_gaap_tags)
            else:
                potential_mappings = [match_non_gaap(non_gaap) for non_gaap in non_us_gaap_tags]
            potential_mappings = [x for x in potential_mappings if x[1] and x[2] > 0.7]
            potential_mappings.sort(key=lambda x: x[2], reverse=True)
            top_mappings = potential_mappings[:20]
            for non_gaap_str, best_usgaap, score in top_mappings:
                concept_mapping[non_gaap_str] = best_usgaap
        
        self._fuzzy_match_cache[cache_key] = concept_mapping.copy()
        return concept_mapping

    def get_quarter_configurations(self) -> Dict[str, Dict[str, int]]:
        """
        Get predefined quarter configurations for different analysis periods.
        
        Returns:
            Dict[str, Dict[str, int]]: Dictionary of configuration names to their details
        """
        return {
            "short_term": {
                "quarters": 4,
                "years": 1,
                "description": "1 year of data - good for recent performance analysis"
            },
            "medium_term": {
                "quarters": 8,
                "years": 2,
                "description": "2 years of data - balanced view for most analyses"
            },
            "long_term": {
                "quarters": 12,
                "years": 3,
                "description": "3 years of data - good for trend analysis"
            },
            "extended": {
                "quarters": 16,
                "years": 4,
                "description": "4 years of data - comprehensive historical view"
            },
            "maximum": {
                "quarters": 20,
                "years": 5,
                "description": "5 years of data - maximum recommended for performance"
            }
        }
    
    def print_quarter_configurations(self):
        """
        Print available quarter configurations with descriptions.
        """
        configs = self.get_quarter_configurations()
        print("\nAvailable Quarter Configurations:")
        print("=" * 60)
        for name, details in configs.items():
            print(f"{name:12} : {details['quarters']:2d} quarters ({details['years']} years) - {details['description']}")
        print("=" * 60)
        print("Usage: create_financial_model(ticker, quarters=configs['medium_term']['quarters'])")

    def _validate_financial_data(self, concept_name: str, value: float, unit: str = None) -> bool:
        """
        Validate that a data point represents actual financial amounts and not other types of data.
        
        Args:
            concept_name: Name of the financial concept
            value: The numeric value to validate
            unit: The unit of measurement (if available)
            
        Returns:
            bool: True if the data appears to be valid financial data, False otherwise
        """
        # Convert to string for pattern matching
        value_str = str(value).lower()
        concept_lower = concept_name.lower()
        
        # 1. Check for year/date patterns
        year_patterns = [
            r'^20\d{2}$',  # Years like 2023, 2024
            r'^19\d{2}$',  # Years like 1999, 2000
            r'^\d{4}$'     # Any 4-digit number that could be a year
        ]
        
        for pattern in year_patterns:
            if re.match(pattern, value_str):
                return False
        
        # 2. Check for employee counts and headcount data
        employee_indicators = [
            'employee', 'headcount', 'personnel', 'staff', 'workforce',
            'full-time', 'part-time', 'fte', 'head count'
        ]
        
        for indicator in employee_indicators:
            if indicator in concept_lower:
                return False
        
        # 3. Check for percentage values (should be excluded from financial statements)
        percentage_indicators = [
            'percentage', 'percent', 'rate', 'ratio', 'margin', 'pct',
            'growth rate', 'return on', 'roi', 'roe', 'roa'
        ]
        
        for indicator in percentage_indicators:
            if indicator in concept_lower:
                return False
        
        # 4. Check for unit-based validation
        if unit:
            unit_lower = unit.lower()
            # Reject non-monetary units
            non_monetary_units = [
                'shares', 'units', 'employees', 'people', 'customers',
                'locations', 'stores', 'facilities', 'countries',
                'percent', 'percentage', 'ratio', 'times', 'days'
            ]
            
            for non_unit in non_monetary_units:
                if non_unit in unit_lower:
                    return False
        
        # 5. Check for suspicious value ranges
        # Reject very small numbers that are likely percentages (0.01 to 0.99)
        if 0.01 <= abs(value) <= 0.99:
            # But allow if it's clearly a financial amount (millions, billions)
            if abs(value) >= 1000000:  # Allow millions+
                pass
            else:
                # Check if concept suggests it should be a percentage
                percentage_concepts = ['margin', 'ratio', 'rate', 'return', 'growth']
                if any(pc in concept_lower for pc in percentage_concepts):
                    return False
        
        # 6. Check for count-based concepts
        count_indicators = [
            'number of', 'count of', 'total number', 'quantity',
            'shares outstanding', 'common stock', 'preferred stock'
        ]
        
        for indicator in count_indicators:
            if indicator in concept_lower:
                # Allow share counts but reject other counts
                if 'share' in indicator and ('outstanding' in indicator or 'stock' in indicator):
                    pass  # Allow share counts
                else:
                    return False
        
        # 7. Check for text-like values
        if isinstance(value, str) and not value.replace('.', '').replace('-', '').isdigit():
            return False
        
        # 8. Check for extreme values that might be errors
        if abs(value) > 1e15:  # Values over 1 quadrillion are suspicious
            return False
        
        # 9. Check for common non-financial concepts
        non_financial_indicators = [
            'age', 'duration', 'length', 'width', 'height', 'weight',
            'temperature', 'speed', 'distance', 'area', 'volume',
            'efficiency', 'productivity', 'satisfaction', 'score'
        ]
        
        for indicator in non_financial_indicators:
            if indicator in concept_lower:
                return False
        
        return True

    def _classify_financial_line_item(self, line_item: str, available_concepts: List[str] = None) -> Dict[str, float]:
        """
        Classify a financial line item into the appropriate statement category using rules-based logic and fuzzy matching.
        
        Args:
            line_item (str): The line item name to classify
            available_concepts (List[str]): List of available XBRL concepts for validation
            
        Returns:
            Dict[str, float]: Dictionary with statement type and confidence score
                {
                    'statement': 'income_statement' | 'balance_sheet' | 'cash_flow',
                    'confidence': 0.0-1.0,
                    'category': 'revenue' | 'expense' | 'asset' | 'liability' | 'equity' | 'operating' | 'investing' | 'financing'
                }
        """
        import re
        from difflib import SequenceMatcher
        
        # Normalize the line item for better matching
        normalized_item = line_item.lower().strip()
        
        # Define comprehensive classification rules based on GAAP standards
        
        # INCOME STATEMENT CLASSIFICATION RULES
        income_statement_patterns = {
            # Revenue patterns (high confidence)
            'revenue': {
                'patterns': [
                    r'revenue', r'sales', r'income.*contract', r'fee.*income', 
                    r'commission.*income', r'royalty.*income', r'license.*income',
                    r'product.*revenue', r'service.*revenue', r'net.*sales'
                ],
                'confidence': 0.95,
                'category': 'revenue'
            },
            
            # Cost of goods sold patterns (high confidence)
            'cost_of_goods': {
                'patterns': [
                    r'cost.*goods', r'cost.*revenue', r'cost.*sales', r'cost.*services',
                    r'cost.*products', r'cost.*contract', r'cost.*revenue.*contract'
                ],
                'confidence': 0.95,
                'category': 'expense'
            },
            
            # Operating expense patterns (high confidence)
            'operating_expenses': {
                'patterns': [
                    r'research.*development', r'rd.*expense', r'selling.*general.*administrative',
                    r'sga', r'operating.*expense', r'administrative.*expense',
                    r'marketing.*expense', r'advertising.*expense', r'promotion.*expense'
                ],
                'confidence': 0.90,
                'category': 'expense'
            },
            
            # Depreciation and amortization (high confidence)
            'depreciation': {
                'patterns': [
                    r'depreciation', r'amortization', r'depreciation.*amortization',
                    r'accumulated.*depreciation', r'accumulated.*amortization'
                ],
                'confidence': 0.90,
                'category': 'expense'
            },
            
            # Stock-based compensation (high confidence)
            'stock_compensation': {
                'patterns': [
                    r'stock.*based.*compensation', r'stock.*compensation', r'equity.*compensation',
                    r'option.*expense', r'warrant.*expense', r'rsu.*expense'
                ],
                'confidence': 0.90,
                'category': 'expense'
            },
            
            # Restructuring and impairment (high confidence)
            'restructuring': {
                'patterns': [
                    r'restructuring', r'impairment', r'write.*off', r'write.*down',
                    r'goodwill.*impairment', r'intangible.*impairment'
                ],
                'confidence': 0.90,
                'category': 'expense'
            },
            
            # Operating income patterns (high confidence)
            'operating_income': {
                'patterns': [
                    r'operating.*income', r'operating.*profit', r'operating.*loss',
                    r'ebit', r'earnings.*before.*interest.*tax'
                ],
                'confidence': 0.95,
                'category': 'income'
            },
            
            # Interest patterns (high confidence)
            'interest': {
                'patterns': [
                    r'interest.*expense', r'interest.*income', r'interest.*cost',
                    r'interest.*revenue', r'interest.*paid', r'interest.*received'
                ],
                'confidence': 0.90,
                'category': 'expense'
            },
            
            # Tax patterns (high confidence)
            'taxes': {
                'patterns': [
                    r'income.*tax', r'tax.*expense', r'tax.*benefit', r'provision.*tax',
                    r'deferred.*tax', r'current.*tax', r'effective.*tax.*rate'
                ],
                'confidence': 0.95,
                'category': 'expense'
            },
            
            # Net income patterns (highest confidence)
            'net_income': {
                'patterns': [
                    r'net.*income', r'net.*loss', r'net.*earnings', r'net.*profit',
                    r'net.*income.*loss', r'comprehensive.*income', r'net.*income.*attributable'
                ],
                'confidence': 0.98,
                'category': 'income'
            },
            
            # Earnings per share patterns (high confidence)
            'eps': {
                'patterns': [
                    r'earnings.*per.*share', r'eps', r'basic.*eps', r'diluted.*eps',
                    r'weighted.*average.*shares', r'shares.*outstanding'
                ],
                'confidence': 0.90,
                'category': 'metric'
            }
        }
        
        # BALANCE SHEET CLASSIFICATION RULES
        balance_sheet_patterns = {
            # Current assets patterns (high confidence)
            'current_assets': {
                'patterns': [
                    r'cash.*equivalent', r'cash.*equivalents', r'cash.*restricted',
                    r'accounts.*receivable', r'inventory', r'prepaid.*expense',
                    r'short.*term.*investment', r'marketable.*security',
                    r'current.*asset', r'asset.*current'
                ],
                'confidence': 0.95,
                'category': 'asset'
            },
            
            # Non-current assets patterns (high confidence)
            'non_current_assets': {
                'patterns': [
                    r'property.*plant.*equipment', r'ppe', r'fixed.*asset',
                    r'goodwill', r'intangible.*asset', r'long.*term.*investment',
                    r'deferred.*tax.*asset', r'other.*asset.*noncurrent',
                    r'noncurrent.*asset', r'asset.*noncurrent'
                ],
                'confidence': 0.95,
                'category': 'asset'
            },
            
            # Total assets patterns (highest confidence)
            'total_assets': {
                'patterns': [
                    r'total.*asset', r'asset.*total', r'consolidated.*asset'
                ],
                'confidence': 0.98,
                'category': 'asset'
            },
            
            # Current liabilities patterns (high confidence)
            'current_liabilities': {
                'patterns': [
                    r'accounts.*payable', r'accrued.*liability', r'accrued.*expense',
                    r'deferred.*revenue', r'short.*term.*debt', r'current.*liability',
                    r'liability.*current', r'note.*payable.*current'
                ],
                'confidence': 0.95,
                'category': 'liability'
            },
            
            # Non-current liabilities patterns (high confidence)
            'non_current_liabilities': {
                'patterns': [
                    r'long.*term.*debt', r'deferred.*tax.*liability',
                    r'other.*liability.*noncurrent', r'noncurrent.*liability',
                    r'liability.*noncurrent', r'note.*payable.*noncurrent'
                ],
                'confidence': 0.95,
                'category': 'liability'
            },
            
            # Total liabilities patterns (highest confidence)
            'total_liabilities': {
                'patterns': [
                    r'total.*liability', r'liability.*total', r'consolidated.*liability'
                ],
                'confidence': 0.98,
                'category': 'liability'
            },
            
            # Equity patterns (high confidence)
            'equity': {
                'patterns': [
                    r'common.*stock', r'preferred.*stock', r'additional.*paid.*capital',
                    r'retained.*earnings', r'treasury.*stock', r'accumulated.*other.*comprehensive',
                    r'stockholders.*equity', r'shareholders.*equity', r'equity.*total',
                    r'paid.*capital', r'capital.*stock'
                ],
                'confidence': 0.95,
                'category': 'equity'
            },
            
            # Working capital patterns (high confidence)
            'working_capital': {
                'patterns': [
                    r'working.*capital', r'net.*working.*capital'
                ],
                'confidence': 0.90,
                'category': 'metric'
            }
        }
        
        # CASH FLOW STATEMENT CLASSIFICATION RULES
        cash_flow_patterns = {
            # Operating activities patterns (high confidence)
            'operating_activities': {
                'patterns': [
                    r'net.*cash.*operating', r'cash.*operating.*activity',
                    r'operating.*cash.*flow', r'cash.*provided.*operating',
                    r'cash.*used.*operating', r'net.*cash.*provided.*operating',
                    r'operating.*activity', r'cash.*flow.*operating'
                ],
                'confidence': 0.95,
                'category': 'operating'
            },
            
            # Operating adjustments patterns (high confidence)
            'operating_adjustments': {
                'patterns': [
                    r'depreciation.*amortization', r'stock.*based.*compensation',
                    r'deferred.*income.*tax', r'deferred.*tax', r'provision.*bad.*debt',
                    r'gain.*loss.*sale.*asset', r'impairment.*charge'
                ],
                'confidence': 0.90,
                'category': 'operating'
            },
            
            # Working capital changes patterns (high confidence)
            'working_capital_changes': {
                'patterns': [
                    r'change.*account.*receivable', r'change.*inventory',
                    r'change.*account.*payable', r'change.*deferred.*revenue',
                    r'increase.*decrease.*receivable', r'increase.*decrease.*inventory',
                    r'increase.*decrease.*payable', r'change.*working.*capital'
                ],
                'confidence': 0.90,
                'category': 'operating'
            },
            
            # Investing activities patterns (high confidence)
            'investing_activities': {
                'patterns': [
                    r'net.*cash.*investing', r'cash.*investing.*activity',
                    r'investing.*cash.*flow', r'cash.*provided.*investing',
                    r'cash.*used.*investing', r'net.*cash.*provided.*investing',
                    r'investing.*activity', r'cash.*flow.*investing'
                ],
                'confidence': 0.95,
                'category': 'investing'
            },
            
            # Capital expenditures patterns (high confidence)
            'capital_expenditures': {
                'patterns': [
                    r'capital.*expenditure', r'capex', r'payment.*property.*plant.*equipment',
                    r'payment.*ppe', r'acquisition.*property.*plant.*equipment',
                    r'purchase.*property.*plant.*equipment'
                ],
                'confidence': 0.95,
                'category': 'investing'
            },
            
            # Acquisitions and investments patterns (high confidence)
            'acquisitions_investments': {
                'patterns': [
                    r'acquisition.*business', r'payment.*acquire.*business',
                    r'payment.*investment', r'proceeds.*sale.*investment',
                    r'proceeds.*maturity.*investment', r'purchase.*investment',
                    r'sale.*investment'
                ],
                'confidence': 0.90,
                'category': 'investing'
            },
            
            # Financing activities patterns (high confidence)
            'financing_activities': {
                'patterns': [
                    r'net.*cash.*financing', r'cash.*financing.*activity',
                    r'financing.*cash.*flow', r'cash.*provided.*financing',
                    r'cash.*used.*financing', r'net.*cash.*provided.*financing',
                    r'financing.*activity', r'cash.*flow.*financing'
                ],
                'confidence': 0.95,
                'category': 'financing'
            },
            
            # Debt financing patterns (high confidence)
            'debt_financing': {
                'patterns': [
                    r'proceeds.*debt', r'proceeds.*issuance.*debt',
                    r'repayment.*debt', r'repayment.*long.*term.*debt',
                    r'borrowing.*debt', r'issuance.*debt'
                ],
                'confidence': 0.90,
                'category': 'financing'
            },
            
            # Equity financing patterns (high confidence)
            'equity_financing': {
                'patterns': [
                    r'proceeds.*stock', r'proceeds.*issuance.*stock',
                    r'proceeds.*common.*stock', r'proceeds.*preferred.*stock',
                    r'repurchase.*stock', r'payment.*repurchase.*stock',
                    r'dividend.*paid', r'payment.*dividend'
                ],
                'confidence': 0.90,
                'category': 'financing'
            },
            
            # Net change patterns (highest confidence)
            'net_change': {
                'patterns': [
                    r'net.*change.*cash', r'net.*increase.*decrease.*cash',
                    r'cash.*beginning.*period', r'cash.*end.*period',
                    r'effect.*exchange.*rate', r'exchange.*rate.*effect'
                ],
                'confidence': 0.98,
                'category': 'net_change'
            }
        }
        
        # Combine all patterns
        all_patterns = {
            'income_statement': income_statement_patterns,
            'balance_sheet': balance_sheet_patterns,
            'cash_flow': cash_flow_patterns
        }
        
        # Initialize results
        best_match = {
            'statement': None,
            'confidence': 0.0,
            'category': None
        }
        
        # Check each statement type
        for statement_type, patterns in all_patterns.items():
            for pattern_name, pattern_info in patterns.items():
                for pattern in pattern_info['patterns']:
                    if re.search(pattern, normalized_item, re.IGNORECASE):
                        confidence = pattern_info['confidence']
                        
                        # Boost confidence for exact matches
                        if normalized_item == pattern.lower():
                            confidence = min(1.0, confidence + 0.05)
                        
                        # Boost confidence for partial word matches
                        words = normalized_item.split()
                        pattern_words = pattern.lower().split()
                        word_matches = sum(1 for word in words if any(pword in word for pword in pattern_words))
                        if word_matches > 0:
                            confidence = min(1.0, confidence + (word_matches * 0.02))
                        
                        # Update best match if this is better
                        if confidence > best_match['confidence']:
                            best_match = {
                                'statement': statement_type,
                                'confidence': confidence,
                                'category': pattern_info['category']
                            }
        
        # Additional fuzzy matching for edge cases
        if best_match['confidence'] < 0.7 and available_concepts:
            # Try fuzzy matching with available concepts
            for concept in available_concepts:
                similarity = SequenceMatcher(None, normalized_item, concept.lower()).ratio()
                if similarity > 0.8:
                    # Use the fuzzy match result to help classify
                    fuzzy_result = self._classify_financial_line_item(concept)
                    if fuzzy_result['confidence'] > best_match['confidence']:
                        best_match = {
                            'statement': fuzzy_result['statement'],
                            'confidence': fuzzy_result['confidence'] * 0.9,  # Slightly reduce confidence for fuzzy matches
                            'category': fuzzy_result['category']
                        }
        
        # Fallback rules for common patterns
        if best_match['confidence'] < 0.5:
            # Apply fallback rules
            fallback_rules = [
                # Income statement fallbacks
                (r'expense|cost|charge|loss', 'income_statement', 0.6, 'expense'),
                (r'income|profit|earnings|revenue', 'income_statement', 0.6, 'income'),
                (r'eps|per.*share', 'income_statement', 0.7, 'metric'),
                
                # Balance sheet fallbacks
                (r'asset|receivable|inventory|equipment', 'balance_sheet', 0.6, 'asset'),
                (r'liability|payable|debt|obligation', 'balance_sheet', 0.6, 'liability'),
                (r'equity|stock|capital|retained', 'balance_sheet', 0.6, 'equity'),
                
                # Cash flow fallbacks
                (r'cash.*flow|cash.*provided|cash.*used', 'cash_flow', 0.7, 'operating'),
                (r'change.*cash|increase.*decrease.*cash', 'cash_flow', 0.7, 'net_change'),
            ]
            
            for pattern, statement, confidence, category in fallback_rules:
                if re.search(pattern, normalized_item, re.IGNORECASE):
                    if confidence > best_match['confidence']:
                        best_match = {
                            'statement': statement,
                            'confidence': confidence,
                            'category': category
                        }
        
        return best_match

    def _validate_statement_classification(self, line_item: str, statement_type: str, available_concepts: List[str] = None) -> bool:
        """
        Validate that a line item is correctly classified into its statement type.
        
        Args:
            line_item (str): The line item name
            statement_type (str): The proposed statement type
            available_concepts (List[str]): List of available XBRL concepts for validation
            
        Returns:
            bool: True if classification is valid, False otherwise
        """
        classification = self._classify_financial_line_item(line_item, available_concepts)
        
        # Check if the classification matches the proposed statement type
        if classification['statement'] == statement_type:
            return True
        
        # If confidence is very low, we might want to reclassify
        if classification['confidence'] > 0.8 and classification['statement'] != statement_type:
            return False
        
        # For moderate confidence, check if the proposed type makes sense
        if classification['confidence'] > 0.6:
            # Allow some flexibility for edge cases
            return True
        
        return True  # Default to accepting the classification

    def _auto_classify_and_validate_line_items(self, comprehensive_facts: Dict, progress_callback=None) -> Dict[str, Dict[str, str]]:
        """
        Automatically classify line items into appropriate financial statements using the classification system.
        
        Args:
            comprehensive_facts (Dict): Dictionary of all available XBRL facts
            progress_callback: Optional progress callback function
            
        Returns:
            Dict[str, Dict[str, str]]: Dictionary mapping statement types to line item classifications
        """
        def progress(msg):
            if progress_callback:
                progress_callback(msg)
            else:
                print(msg)
        
        progress("    • Auto-classifying line items into financial statements...")
        
        # Get all available concepts
        available_concepts = list(comprehensive_facts.keys())
        
        # Initialize classification results
        classifications = {
            'income_statement': {},
            'balance_sheet': {},
            'cash_flow': {}
        }
        
        # Track classification statistics
        stats = {
            'total_items': len(available_concepts),
            'classified_items': 0,
            'high_confidence': 0,
            'medium_confidence': 0,
            'low_confidence': 0,
            'unclassified': 0
        }
        
        # Classify each concept
        for concept in available_concepts:
            # Extract the concept name (remove namespace prefix)
            concept_name = concept.split(':')[-1] if ':' in concept else concept
            
            # Classify the line item
            classification = self._classify_financial_line_item(concept_name, available_concepts)
            
            if classification['statement']:
                statement_type = classification['statement']
                confidence = classification['confidence']
                category = classification['category']
                
                # Store the classification
                classifications[statement_type][concept] = {
                    'confidence': confidence,
                    'category': category,
                    'concept_name': concept_name
                }
                
                stats['classified_items'] += 1
                
                # Track confidence levels
                if confidence >= 0.8:
                    stats['high_confidence'] += 1
                elif confidence >= 0.6:
                    stats['medium_confidence'] += 1
                else:
                    stats['low_confidence'] += 1
            else:
                stats['unclassified'] += 1
        
        # Report classification statistics
        progress(f"    • Classification complete:")
        progress(f"      - Total items: {stats['total_items']}")
        progress(f"      - Classified: {stats['classified_items']} ({stats['classified_items']/stats['total_items']*100:.1f}%)")
        progress(f"      - High confidence (≥80%): {stats['high_confidence']}")
        progress(f"      - Medium confidence (60-79%): {stats['medium_confidence']}")
        progress(f"      - Low confidence (<60%): {stats['low_confidence']}")
        progress(f"      - Unclassified: {stats['unclassified']}")
        
        # Report breakdown by statement
        for statement_type, items in classifications.items():
            if items:
                high_conf = sum(1 for item in items.values() if item['confidence'] >= 0.8)
                progress(f"      - {statement_type.replace('_', ' ').title()}: {len(items)} items ({high_conf} high confidence)")
        
        return classifications

    def _create_improved_financial_model(self, comprehensive_facts: Dict, annual_data: Dict, quarterly_data: Dict, 
                                       enhanced_fuzzy_matching: bool = True, progress_callback=None) -> Dict[str, pd.DataFrame]:
        """
        Create an improved financial model using the new classification system.
        
        Args:
            comprehensive_facts (Dict): All available XBRL facts
            annual_data (Dict): Annual data points
            quarterly_data (Dict): Quarterly data points
            enhanced_fuzzy_matching (bool): Whether to use enhanced fuzzy matching
            progress_callback: Optional progress callback function
            
        Returns:
            Dict[str, pd.DataFrame]: Improved financial model with better classification
        """
        def progress(msg):
            if progress_callback:
                progress_callback(msg)
            else:
                print(msg)
        
        progress("    • Creating improved financial model with enhanced classification...")
        
        # Auto-classify all line items
        classifications = self._auto_classify_and_validate_line_items(comprehensive_facts, progress_callback)
        
        # Initialize financial model
        financial_model = {
            'annual_income_statement': pd.DataFrame(),
            'annual_balance_sheet': pd.DataFrame(),
            'annual_cash_flow': pd.DataFrame(),
            'quarterly_income_statement': pd.DataFrame(),
            'quarterly_balance_sheet': pd.DataFrame(),
            'quarterly_cash_flow': pd.DataFrame()
        }
        
        # Process each statement type
        for statement_type, classified_items in classifications.items():
            if not classified_items:
                continue
                
            progress(f"    • Processing {statement_type.replace('_', ' ')} items...")
            
            # Filter items by confidence level (focus on high and medium confidence)
            high_confidence_items = {k: v for k, v in classified_items.items() if v['confidence'] >= 0.6}
            
            if not high_confidence_items:
                progress(f"      ⚠ No high-confidence items found for {statement_type}")
                continue
            
            # Create mapping for this statement type
            statement_mapping = {}
            for concept, info in high_confidence_items.items():
                concept_name = info['concept_name']
                # Use the classified concept name as the key
                statement_mapping[concept_name] = concept
            
            progress(f"      ✓ Found {len(high_confidence_items)} high-confidence items")
            
            # Create DataFrames for annual and quarterly data
            for period in ['annual', 'quarterly']:
                data_source = annual_data if period == 'annual' else quarterly_data
                
                # Extract data for this statement type
                statement_data = {}
                for concept, info in high_confidence_items.items():
                    if concept in data_source:
                        # Group data by date
                        for data_point in data_source[concept]:
                            date = data_point.get('end', '')
                            value = data_point.get('val', 0)
                            
                            if date not in statement_data:
                                statement_data[date] = {}
                            
                            concept_name = info['concept_name']
                            statement_data[date][concept_name] = value
                
                if statement_data:
                    # Create DataFrame
                    df = pd.DataFrame.from_dict(statement_data, orient='index')
                    df.index = pd.to_datetime(df.index)
                    df = df.sort_index()
                    
                    # Store in financial model
                    key = f'{period}_{statement_type}'
                    financial_model[key] = df
                    
                    progress(f"      ✓ {period.title()}: {len(df)} periods, {len(df.columns)} line items")
                else:
                    progress(f"      ⚠ No {period} data found for {statement_type}")
        
        # Validate the financial model
        progress("    • Validating financial model...")
        validation_results = self._validate_financial_model_consistency(financial_model)
        
        for statement_type, result in validation_results.items():
            if result['valid']:
                progress(f"      ✓ {statement_type}: Valid ({result['data_points']} data points)")
            else:
                progress(f"      ⚠ {statement_type}: {result['issues']}")
        
        progress("    ✓ Improved financial model created successfully!")
        return financial_model

    def _validate_financial_model_consistency(self, financial_model: Dict[str, pd.DataFrame]) -> Dict[str, Dict]:
        """
        Validate the consistency and quality of the financial model.
        
        Args:
            financial_model (Dict[str, pd.DataFrame]): The financial model to validate
            
        Returns:
            Dict[str, Dict]: Validation results for each statement type
        """
        validation_results = {}
        
        for statement_type, df in financial_model.items():
            result = {
                'valid': False,
                'data_points': 0,
                'issues': []
            }
            
            if df.empty:
                result['issues'].append("No data available")
            else:
                result['data_points'] = len(df) * len(df.columns)
                
                # Check for basic data quality
                if df.isnull().all().all():
                    result['issues'].append("All data is null")
                elif df.isnull().sum().sum() > len(df) * len(df.columns) * 0.5:
                    result['issues'].append("More than 50% of data is null")
                else:
                    result['valid'] = True
                    
                    # Check for reasonable value ranges
                    numeric_cols = df.select_dtypes(include=[np.number]).columns
                    for col in numeric_cols:
                        if df[col].abs().max() > 1e15:  # Very large numbers
                            result['issues'].append(f"Unusually large values in {col}")
                        elif df[col].abs().min() > 0 and df[col].abs().min() < 1e-10:  # Very small numbers
                            result['issues'].append(f"Unusually small values in {col}")
            
            validation_results[statement_type] = result
        
        return validation_results

    def _create_user_friendly_title_mapping(self) -> Dict[str, str]:
        """
        Create a mapping from technical GAAP concept names to user-friendly, descriptive titles.
        
        Returns:
            Dict[str, str]: Mapping from concept names to user-friendly titles
        """
        return {
            # Income Statement - Revenue Section
            'RevenueFromContractWithCustomerExcludingAssessedTax': 'Revenue',
            'Revenues': 'Revenue',
            'SalesRevenueNet': 'Revenue',
            'RevenueFromContractWithCustomer': 'Revenue',
            'SalesRevenueGoodsNet': 'Product Revenue',
            'SalesRevenueServicesNet': 'Service Revenue',
            'SalesRevenueNetOfReturnsAndAllowances': 'Net Sales',
            
            # New required line items
            'Revenue Stream One': 'Revenue Stream One',
            'Revenue Stream Two': 'Revenue Stream Two',
            'Total Revenues': 'Total Revenues',
            'OpEx': 'OpEx',
            'EBITDA': 'EBITDA',
            
            # Cost of Revenue
            'CostOfRevenue': 'Cost of Revenue',
            'CostOfGoodsAndServicesSold': 'Cost of Goods Sold',
            'CostOfGoodsSold': 'Cost of Goods Sold',
            'CostOfRevenueExcludingDepreciationAndAmortization': 'Cost of Revenue (ex. D&A)',
            
            # Gross Profit
            'GrossProfit': 'Gross Profit',
            'GrossProfitLoss': 'Gross Profit',
            
            # Operating Expenses
            'ResearchAndDevelopmentExpense': 'Research & Development',
            'SellingGeneralAndAdministrativeExpense': 'Selling, General & Administrative',
            'SellingAndMarketingExpense': 'Selling & Marketing',
            'GeneralAndAdministrativeExpense': 'General & Administrative',
            'MarketingExpense': 'Marketing',
            'AdvertisingExpense': 'Advertising',
            'PromotionalExpense': 'Promotional',
            
            # SG&A and Misc
            'SG&A': 'SG&A',
            'Misc': 'Misc',
            'Total OpEx': 'Total OpEx',
            
            # Depreciation and Amortization
            'DepreciationAndAmortization': 'Depreciation & Amortization',
            'Depreciation': 'Depreciation',
            'AmortizationOfIntangibleAssets': 'Amortization',
            'AmortizationOfDeferredCharges': 'Amortization of Deferred Charges',
            
            # Stock-based Compensation
            'StockBasedCompensationExpense': 'Stock-Based Compensation',
            'CompensationAndRelatedExpense': 'Compensation Expense',
            'EmployeeCompensationExpense': 'Employee Compensation',
            
            # Restructuring and Impairment
            'RestructuringCharges': 'Restructuring Charges',
            'ImpairmentCharges': 'Impairment Charges',
            'GoodwillImpairmentLoss': 'Goodwill Impairment',
            'IntangibleAssetsImpairmentLoss': 'Intangible Asset Impairment',
            'AssetImpairmentCharges': 'Asset Impairment',
            
            # Operating Income
            'OperatingIncomeLoss': 'Operating Income',
            'OperatingExpenses': 'Operating Expenses',
            'OtherOperatingExpenses': 'Other Operating Expenses',
            'OtherOperatingIncomeExpense': 'Other Operating Income/(Expense)',
            
            # Non-Operating Items
            'InterestExpense': 'Interest Expense',
            'InterestIncome': 'Interest Income',
            'InterestExpenseDebt': 'Interest Expense on Debt',
            'InterestExpenseCapitalized': 'Capitalized Interest',
            'InterestIncomeExpenseNet': 'Net Interest Income/(Expense)',
            
            # Gains and Losses
            'GainLossOnSaleOfAssets': 'Gain/(Loss) on Asset Sales',
            'GainLossOnSaleOfPropertyPlantEquipment': 'Gain/(Loss) on PP&E Sales',
            'GainLossOnSaleOfInvestments': 'Gain/(Loss) on Investment Sales',
            'GainLossOnSaleOfBusiness': 'Gain/(Loss) on Business Sales',
            
            # Foreign Currency
            'ForeignCurrencyGainLoss': 'Foreign Currency Gain/(Loss)',
            'ForeignCurrencyTransactionGainLoss': 'Foreign Currency Transaction Gain/(Loss)',
            
            # Other Income/Expense
            'OtherIncomeExpenseNet': 'Other Income/(Expense)',
            'OtherNonoperatingIncomeExpense': 'Other Non-Operating Income/(Expense)',
            'MiscellaneousIncomeExpense': 'Miscellaneous Income/(Expense)',
            
            # Income Before Taxes
            'IncomeLossFromContinuingOperationsBeforeIncomeTaxes': 'Income Before Taxes',
            'IncomeLossFromContinuingOperationsBeforeIncomeTaxesDomestic': 'Income Before Taxes (Domestic)',
            'IncomeLossFromContinuingOperationsBeforeIncomeTaxesForeign': 'Income Before Taxes (Foreign)',
            
            # Income Taxes
            'IncomeTaxExpenseBenefit': 'Income Tax Expense',
            'ProvisionForIncomeTaxes': 'Income Tax Provision',
            'CurrentIncomeTaxExpense': 'Current Income Tax',
            'DeferredIncomeTaxExpense': 'Deferred Income Tax',
            'EffectiveIncomeTaxRateReconciliation': 'Effective Tax Rate',
            'Tax Expense': 'Tax Expense',
            
            # Net Income
            'NetIncomeLoss': 'Net Income',
            'NetIncomeLossAvailableToCommonStockholdersBasic': 'Net Income Available to Common',
            'NetIncomeLossAttributableToParent': 'Net Income Attributable to Parent',
            'NetIncomeLossAttributableToNoncontrollingInterest': 'Net Income Attributable to Noncontrolling Interest',
            'Net Income': 'Net Income',
            
            # Earnings Per Share
            'EarningsPerShareBasic': 'EPS (Basic)',
            'EarningsPerShareDiluted': 'EPS (Diluted)',
            'WeightedAverageNumberOfSharesBasic': 'Weighted Average Shares (Basic)',
            'WeightedAverageNumberOfSharesOutstandingDiluted': 'Weighted Average Shares (Diluted)',
            
            # Balance Sheet - Current Assets
            'CashAndCashEquivalentsAtCarryingValue': 'Cash & Cash Equivalents',
            'Cash': 'Cash',
            'CashEquivalentsAtCarryingValue': 'Cash Equivalents',
            'RestrictedCash': 'Restricted Cash',
            'ShortTermInvestments': 'Short-Term Investments',
            'MarketableSecurities': 'Marketable Securities',
            'AccountsReceivableNet': 'Accounts Receivable',
            'AccountsReceivableGross': 'Gross Accounts Receivable',
            'AllowanceForDoubtfulAccountsReceivable': 'Allowance for Doubtful Accounts',
            'InventoryNet': 'Inventory',
            'InventoryGross': 'Gross Inventory',
            'InventoryReserves': 'Inventory Reserves',
            'PrepaidExpensesAndOtherCurrentAssets': 'Prepaid Expenses & Other Current Assets',
            'PrepaidExpense': 'Prepaid Expenses',
            'OtherAssetsCurrent': 'Other Current Assets',
            
            # Non-Current Assets
            'PropertyPlantAndEquipmentNet': 'Property, Plant & Equipment (Net)',
            'PropertyPlantAndEquipmentGross': 'Property, Plant & Equipment (Gross)',
            'AccumulatedDepreciationDepletionAndAmortizationPropertyPlantAndEquipment': 'Accumulated Depreciation',
            'LandAndBuildings': 'Land & Buildings',
            'MachineryAndEquipment': 'Machinery & Equipment',
            'ComputerSoftwareAndEquipment': 'Computer Software & Equipment',
            'ConstructionInProgress': 'Construction in Progress',
            'Goodwill': 'Goodwill',
            'IntangibleAssetsNet': 'Intangible Assets (Net)',
            'IntangibleAssetsGross': 'Intangible Assets (Gross)',
            'AccumulatedAmortizationOfIntangibleAssets': 'Accumulated Amortization',
            'LongTermInvestments': 'Long-Term Investments',
            'InvestmentsInAffiliates': 'Investments in Affiliates',
            'DeferredTaxAssetsNet': 'Deferred Tax Assets',
            'DeferredTaxAssetsGross': 'Deferred Tax Assets (Gross)',
            'OtherAssetsNoncurrent': 'Other Non-Current Assets',
            
            # Current Liabilities
            'AccountsPayable': 'Accounts Payable',
            'AccountsPayableCurrent': 'Accounts Payable',
            'AccruedLiabilitiesCurrent': 'Accrued Liabilities',
            'AccruedExpenses': 'Accrued Expenses',
            'AccruedCompensationAndBenefits': 'Accrued Compensation & Benefits',
            'AccruedIncomeTaxesCurrent': 'Accrued Income Taxes',
            'DeferredRevenueCurrent': 'Deferred Revenue',
            'ContractWithCustomerLiability': 'Contract Liability',
            'ShortTermDebt': 'Short-Term Debt',
            'CommercialPaper': 'Commercial Paper',
            'CurrentMaturitiesOfLongTermDebt': 'Current Portion of Long-Term Debt',
            'OtherLiabilitiesCurrent': 'Other Current Liabilities',
            
            # Non-Current Liabilities
            'LongTermDebt': 'Long-Term Debt',
            'LongTermDebtNoncurrent': 'Long-Term Debt',
            'NotesPayable': 'Notes Payable',
            'BondsPayable': 'Bonds Payable',
            'DeferredTaxLiabilitiesNet': 'Deferred Tax Liabilities',
            'DeferredTaxLiabilitiesGross': 'Deferred Tax Liabilities (Gross)',
            'DeferredRevenueNoncurrent': 'Deferred Revenue (Non-Current)',
            'OtherLiabilitiesNoncurrent': 'Other Non-Current Liabilities',
            
            # Stockholders' Equity
            'CommonStockValue': 'Common Stock',
            'CommonStockSharesAuthorized': 'Common Stock (Authorized)',
            'CommonStockSharesIssued': 'Common Stock (Issued)',
            'CommonStockSharesOutstanding': 'Common Stock (Outstanding)',
            'PreferredStockValue': 'Preferred Stock',
            'PreferredStockSharesAuthorized': 'Preferred Stock (Authorized)',
            'PreferredStockSharesIssued': 'Preferred Stock (Issued)',
            'AdditionalPaidInCapital': 'Additional Paid-In Capital',
            'PaidInCapital': 'Paid-In Capital',
            'RetainedEarningsAccumulatedDeficit': 'Retained Earnings',
            'RetainedEarnings': 'Retained Earnings',
            'AccumulatedDeficit': 'Accumulated Deficit',
            'AccumulatedOtherComprehensiveIncomeLossNetOfTax': 'Accumulated Other Comprehensive Income',
            'TreasuryStockValue': 'Treasury Stock',
            'TreasuryStockShares': 'Treasury Stock (Shares)',
            'StockholdersEquity': 'Stockholders\' Equity',
            'StockholdersEquityIncludingPortionAttributableToNoncontrollingInterest': 'Total Stockholders\' Equity',
            
            # Cash Flow Statement - Operating Activities
            'NetCashProvidedByUsedInOperatingActivities': 'Net Cash from Operating Activities',
            'NetCashProvidedByOperatingActivities': 'Net Cash from Operating Activities',
            'NetCashUsedInOperatingActivities': 'Net Cash Used in Operating Activities',
            'CashAndCashEquivalentsPeriodIncreaseDecrease': 'Net Change in Cash',
            'NetIncomeLoss': 'Net Income',
            'DepreciationAndAmortization': 'Depreciation & Amortization',
            'Depreciation': 'Depreciation',
            'AmortizationOfIntangibleAssets': 'Amortization',
            'StockBasedCompensationExpense': 'Stock-Based Compensation',
            'DeferredIncomeTaxExpenseBenefit': 'Deferred Income Taxes',
            'DeferredTaxExpenseBenefit': 'Deferred Taxes',
            'ProvisionForDoubtfulAccounts': 'Provision for Doubtful Accounts',
            'GainLossOnSaleOfAssets': 'Gain/(Loss) on Asset Sales',
            'GainLossOnSaleOfPropertyPlantEquipment': 'Gain/(Loss) on PP&E Sales',
            'ImpairmentCharges': 'Impairment Charges',
            'RestructuringCharges': 'Restructuring Charges',
            
            # Working Capital Changes
            'IncreaseDecreaseInAccountsReceivable': 'Change in Accounts Receivable',
            'IncreaseDecreaseInInventory': 'Change in Inventory',
            'IncreaseDecreaseInPrepaidExpenses': 'Change in Prepaid Expenses',
            'IncreaseDecreaseInAccountsPayable': 'Change in Accounts Payable',
            'IncreaseDecreaseInAccruedLiabilities': 'Change in Accrued Liabilities',
            'IncreaseDecreaseInDeferredRevenue': 'Change in Deferred Revenue',
            'IncreaseDecreaseInOtherWorkingCapital': 'Change in Other Working Capital',
            'NetCashProvidedByUsedInOperatingActivitiesContinuingOperations': 'Net Cash from Operating Activities (Continuing)',
            
            # Investing Activities
            'NetCashProvidedByUsedInInvestingActivities': 'Net Cash from Investing Activities',
            'NetCashProvidedByInvestingActivities': 'Net Cash from Investing Activities',
            'NetCashUsedInInvestingActivities': 'Net Cash Used in Investing Activities',
            'PaymentsToAcquirePropertyPlantAndEquipment': 'Capital Expenditures',
            'CapitalExpenditures': 'Capital Expenditures',
            'PaymentsToAcquireBusinessesNetOfCashAcquired': 'Acquisitions (Net of Cash)',
            'PaymentsToAcquireBusinesses': 'Acquisitions',
            'ProceedsFromSaleOfBusinesses': 'Proceeds from Business Sales',
            'PaymentsToAcquireInvestments': 'Payments for Investments',
            'ProceedsFromSaleOfInvestments': 'Proceeds from Investment Sales',
            'ProceedsFromMaturitiesPrepaymentsAndCallsOfAvailableForSaleSecurities': 'Proceeds from Securities Maturity',
            'ProceedsFromSaleOfAvailableForSaleSecurities': 'Proceeds from Securities Sales',
            'PaymentsToAcquireAvailableForSaleSecurities': 'Payments for Securities',
            'ProceedsFromSaleOfPropertyPlantAndEquipment': 'Proceeds from PP&E Sales',
            'ProceedsFromSaleOfDecommissionedProperty': 'Proceeds from Asset Sales',
            'OtherInvestingActivities': 'Other Investing Activities',
            
            # Financing Activities
            'NetCashProvidedByUsedInFinancingActivities': 'Net Cash from Financing Activities',
            'NetCashProvidedByFinancingActivities': 'Net Cash from Financing Activities',
            'NetCashUsedInFinancingActivities': 'Net Cash Used in Financing Activities',
            'ProceedsFromIssuanceOfLongTermDebt': 'Proceeds from Debt Issuance',
            'ProceedsFromIssuanceOfDebt': 'Proceeds from Debt',
            'RepaymentsOfLongTermDebt': 'Repayments of Long-Term Debt',
            'RepaymentsOfDebt': 'Repayments of Debt',
            'ProceedsFromIssuanceOfCommonStock': 'Proceeds from Common Stock Issuance',
            'ProceedsFromIssuanceOfPreferredStock': 'Proceeds from Preferred Stock Issuance',
            'ProceedsFromStockOptionsExercised': 'Proceeds from Stock Options',
            'PaymentsForRepurchaseOfCommonStock': 'Stock Repurchases',
            'PaymentsForRepurchaseOfPreferredStock': 'Preferred Stock Repurchases',
            'DividendsPaid': 'Dividends Paid',
            'DividendsPaidCommon': 'Common Stock Dividends',
            'DividendsPaidPreferred': 'Preferred Stock Dividends',
            'OtherFinancingActivities': 'Other Financing Activities',
            
            # Net Change and Ending Balance
            'EffectOfExchangeRateOnCashAndCashEquivalents': 'Effect of Exchange Rate Changes',
            'CashAndCashEquivalentsAtCarryingValue': 'Cash & Cash Equivalents',
            'CashAndCashEquivalentsPeriodIncreaseDecrease': 'Net Change in Cash',
            'CashAndCashEquivalentsBeginningOfPeriod': 'Cash at Beginning of Period',
            'CashAndCashEquivalentsEndOfPeriod': 'Cash at End of Period',
            
            # Additional Common Patterns
            'Other': 'Other',
            'OtherExpense': 'Other Expense',
            'OtherIncome': 'Other Income',
            'MiscellaneousExpense': 'Miscellaneous Expense',
            'MiscellaneousIncome': 'Miscellaneous Income',
            'NonoperatingIncomeExpense': 'Non-Operating Income/(Expense)',
            'NonoperatingExpense': 'Non-Operating Expense',
            'NonoperatingIncome': 'Non-Operating Income',
            
            # Working Capital and Financial Metrics
            'WorkingCapital': 'Working Capital',
            'NetWorkingCapital': 'Net Working Capital',
            'TotalDebt': 'Total Debt',
            'NetDebt': 'Net Debt',
            'TotalEquity': 'Total Equity',
            'TotalAssets': 'Total Assets',
            'TotalLiabilities': 'Total Liabilities',
            'TotalLiabilitiesAndStockholdersEquity': 'Total Liabilities & Equity',
            
            # Ratios and Metrics
            'CurrentRatio': 'Current Ratio',
            'QuickRatio': 'Quick Ratio',
            'DebtToEquityRatio': 'Debt-to-Equity Ratio',
            'ReturnOnEquity': 'Return on Equity',
            'ReturnOnAssets': 'Return on Assets',
            'GrossMargin': 'Gross Margin',
            'OperatingMargin': 'Operating Margin',
            'NetMargin': 'Net Margin',
            'EBITDA': 'EBITDA',
            'EBIT': 'EBIT',
            'EBITDAR': 'EBITDAR'
        }
    
    def _get_user_friendly_title(self, concept_name: str) -> str:
        """
        Convert a technical GAAP concept name to a user-friendly title.
        
        Args:
            concept_name (str): The technical concept name
            
        Returns:
            str: User-friendly title
        """
        # Get the title mapping
        title_mapping = self._create_user_friendly_title_mapping()
        
        # Try exact match first
        if concept_name in title_mapping:
            return title_mapping[concept_name]
        
        # Try case-insensitive match
        for key, value in title_mapping.items():
            if key.lower() == concept_name.lower():
                return value
        
        # Remove common prefixes and try again
        prefixes_to_remove = ['us-gaap:', 'dei:', 'srt:', 'us-gaap-']
        clean_name = concept_name
        for prefix in prefixes_to_remove:
            if clean_name.lower().startswith(prefix.lower()):
                clean_name = clean_name[len(prefix):]
                break
        
        # Try mapping with cleaned name
        if clean_name in title_mapping:
            return title_mapping[clean_name]
        
        # Try case-insensitive match with cleaned name
        for key, value in title_mapping.items():
            if key.lower() == clean_name.lower():
                return value
        
        # If no exact match, try to create a user-friendly title from the concept name
        return self._create_friendly_title_from_concept(concept_name)
    
    def _create_friendly_title_from_concept(self, concept_name: str) -> str:
        """
        Create a user-friendly title from a concept name when no mapping exists.
        
        Args:
            concept_name (str): The technical concept name
            
        Returns:
            str: User-friendly title
        """
        import re
        
        # Remove common prefixes
        prefixes_to_remove = ['us-gaap:', 'dei:', 'srt:', 'us-gaap-']
        clean_name = concept_name
        for prefix in prefixes_to_remove:
            if clean_name.lower().startswith(prefix.lower()):
                clean_name = clean_name[len(prefix):]
                break
        
        # Convert camelCase to Title Case with spaces
        clean_name = re.sub(r'([a-z])([A-Z])', r'\1 \2', clean_name)
        
        # Handle common abbreviations
        abbreviations = {
            'PPE': 'PP&E',
            'SG&A': 'SG&A',
            'R&D': 'R&D',
            'EPS': 'EPS',
            'EBITDA': 'EBITDA',
            'EBIT': 'EBIT',
            'ROE': 'ROE',
            'ROA': 'ROA',
            'D&A': 'D&A'
        }
        
        for abbr, replacement in abbreviations.items():
            clean_name = clean_name.replace(abbr, replacement)
        
        # Capitalize properly
        clean_name = clean_name.title()
        
        # Handle special cases
        clean_name = clean_name.replace('And', '&')
        clean_name = clean_name.replace('Or', 'or')
        clean_name = clean_name.replace('Of', 'of')
        clean_name = clean_name.replace('In', 'in')
        clean_name = clean_name.replace('On', 'on')
        clean_name = clean_name.replace('At', 'at')
        clean_name = clean_name.replace('To', 'to')
        clean_name = clean_name.replace('For', 'for')
        clean_name = clean_name.replace('From', 'from')
        clean_name = clean_name.replace('With', 'with')
        clean_name = clean_name.replace('Net', 'Net')
        clean_name = clean_name.replace('Gross', 'Gross')
        clean_name = clean_name.replace('Current', 'Current')
        clean_name = clean_name.replace('Noncurrent', 'Non-Current')
        clean_name = clean_name.replace('Longterm', 'Long-Term')
        clean_name = clean_name.replace('Shortterm', 'Short-Term')
        
        # Clean up multiple spaces
        clean_name = re.sub(r'\s+', ' ', clean_name).strip()
        
        return clean_name

    def _create_user_friendly_parent_map(self) -> Dict[str, str]:
        """
        Create a parent-child mapping for indentation using user-friendly titles.
        
        Returns:
            Dict[str, str]: Parent-child mapping with user-friendly titles
        """
        return {
            # Income Statement - Standard format
            'Revenue': None,  # Top level
            'Cost of Revenue': 'Revenue',
            'Cost of Goods Sold': 'Revenue',
            'Gross Profit': None,  # Calculated total
            
            # Operating expenses
            'Research & Development': 'Operating Expenses',
            'Selling, General & Administrative': 'Operating Expenses',
            'Selling & Marketing': 'Operating Expenses',
            'General & Administrative': 'Operating Expenses',
            'Marketing': 'Operating Expenses',
            'Advertising': 'Operating Expenses',
            'Depreciation & Amortization': 'Operating Expenses',
            'Stock-Based Compensation': 'Operating Expenses',
            'Restructuring Charges': 'Operating Expenses',
            'Impairment Charges': 'Operating Expenses',
            'Other Operating Expenses': 'Operating Expenses',
            'Operating Expenses': None,  # Subtotal
            
            'Operating Income': None,  # Total
            
            # Non-operating items
            'Interest Income': 'Non-Operating Income',
            'Interest Expense': 'Non-Operating Income',
            'Gain/(Loss) on Asset Sales': 'Non-Operating Income',
            'Foreign Currency Gain/(Loss)': 'Non-Operating Income',
            'Other Income/(Expense)': 'Non-Operating Income',
            'Non-Operating Income': None,  # Subtotal
            
            'Income Before Taxes': None,  # Total
            'Income Tax Expense': 'Income Before Taxes',
            'Net Income': None,  # Final total
            
            # Earnings per share
            'EPS (Basic)': 'Net Income',
            'EPS (Diluted)': 'Net Income',
            'Weighted Average Shares (Basic)': 'Net Income',
            'Weighted Average Shares (Diluted)': 'Net Income',
            
            # Balance Sheet - Standard format: Assets = Liabilities + Equity
            # Current Assets
            'Cash & Cash Equivalents': 'Current Assets',
            'Cash': 'Current Assets',
            'Cash Equivalents': 'Current Assets',
            'Restricted Cash': 'Current Assets',
            'Short-Term Investments': 'Current Assets',
            'Marketable Securities': 'Current Assets',
            'Accounts Receivable': 'Current Assets',
            'Inventory': 'Current Assets',
            'Prepaid Expenses': 'Current Assets',
            'Other Current Assets': 'Current Assets',
            'Current Assets': None,  # Subtotal
            
            # Non-Current Assets
            'Property, Plant & Equipment (Net)': 'Non-Current Assets',
            'Goodwill': 'Non-Current Assets',
            'Intangible Assets (Net)': 'Non-Current Assets',
            'Long-Term Investments': 'Non-Current Assets',
            'Deferred Tax Assets': 'Non-Current Assets',
            'Other Non-Current Assets': 'Non-Current Assets',
            'Non-Current Assets': None,  # Subtotal
            
            'Total Assets': None,  # Total Assets
            
            # Current Liabilities
            'Accounts Payable': 'Current Liabilities',
            'Accrued Liabilities': 'Current Liabilities',
            'Accrued Expenses': 'Current Liabilities',
            'Deferred Revenue': 'Current Liabilities',
            'Short-Term Debt': 'Current Liabilities',
            'Other Current Liabilities': 'Current Liabilities',
            'Current Liabilities': None,  # Subtotal
            
            # Non-Current Liabilities
            'Long-Term Debt': 'Non-Current Liabilities',
            'Deferred Tax Liabilities': 'Non-Current Liabilities',
            'Other Non-Current Liabilities': 'Non-Current Liabilities',
            'Non-Current Liabilities': None,  # Subtotal
            
            'Total Liabilities': None,  # Total Liabilities
            
            # Stockholders' Equity
            'Common Stock': 'Stockholders\' Equity',
            'Preferred Stock': 'Stockholders\' Equity',
            'Additional Paid-In Capital': 'Stockholders\' Equity',
            'Retained Earnings': 'Stockholders\' Equity',
            'Accumulated Other Comprehensive Income': 'Stockholders\' Equity',
            'Treasury Stock': 'Stockholders\' Equity',
            'Stockholders\' Equity': None,  # Subtotal
            
            # Calculated metrics
            'Working Capital': None,
            'Total Debt': None,
            
            # Cash Flow Statement - Standard format
            # Operating Activities
            'Net Income': None,  # Starting point
            'Depreciation & Amortization': 'Operating Adjustments',
            'Stock-Based Compensation': 'Operating Adjustments',
            'Deferred Income Taxes': 'Operating Adjustments',
            'Operating Adjustments': None,  # Subtotal
            
            # Changes in Working Capital
            'Change in Accounts Receivable': 'Working Capital Changes',
            'Change in Inventory': 'Working Capital Changes',
            'Change in Accounts Payable': 'Working Capital Changes',
            'Change in Deferred Revenue': 'Working Capital Changes',
            'Change in Other Working Capital': 'Working Capital Changes',
            'Working Capital Changes': None,  # Subtotal
            
            'Other Operating Activities': 'Operating Activities',
            'Operating Activities': None,  # Subtotal
            'Net Cash from Operating Activities': None,  # Total
            
            # Investing Activities
            'Capital Expenditures': 'Investing Activities',
            'Acquisitions': 'Investing Activities',
            'Investments': 'Investing Activities',
            'Proceeds from Investment Sales': 'Investing Activities',
            'Proceeds from Asset Sales': 'Investing Activities',
            'Other Investing Activities': 'Investing Activities',
            'Investing Activities': None,  # Subtotal
            'Net Cash from Investing Activities': None,  # Total
            
            # Financing Activities
            'Proceeds from Debt': 'Financing Activities',
            'Repayments of Debt': 'Financing Activities',
            'Dividends Paid': 'Financing Activities',
            'Stock Repurchases': 'Financing Activities',
            'Proceeds from Stock Issuance': 'Financing Activities',
            'Other Financing Activities': 'Financing Activities',
            'Financing Activities': None,  # Subtotal
            'Net Cash from Financing Activities': None,  # Total
            
            # Net Change and Ending Balance
            'Effect of Exchange Rate Changes': None,
            'Net Change in Cash': None,  # Final total
            'Cash at Beginning of Period': None,
            'Cash at End of Period': None,
            
            # Income Statement - New required format
            'Income Statement': None,  # Section heading
            'Revenues': None,  # Subheading title
            'Revenue Stream One': 'Revenues',  # Child of Revenues
            'Revenue Stream Two': 'Revenues',  # Child of Revenues
            'Total Revenues': 'Revenues',  # Subtotal of Revenues
            'OpEx': None,  # Subheading title
            'Cost of Goods Sold': 'OpEx',  # Child of OpEx
            'SG&A': 'OpEx',  # Child of OpEx
            'Misc': 'OpEx',  # Child of OpEx
            'Total OpEx': 'OpEx',  # Subtotal of OpEx
            'EBITDA': None,  # Total Revenues - Total OpEx
            'Depreciation and Amortization': None,
            'Interest Expense': None,
            'Tax Expense': None,
            'Net Income': None,  # EBITDA - Depreciation and Amortization - Interest Expense - Tax Expense
            
            # Legacy mappings for backward compatibility
            'Revenue': None,  # Top level
            'Cost of Revenue': 'Revenue',
            'Gross Profit': None,  # Calculated total
        }

    def _remove_empty_columns(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Remove columns that are completely empty or contain only empty strings.
        
        Args:
            df: DataFrame to clean
            
        Returns:
            DataFrame with empty columns removed
        """
        if df.empty:
            return df
        
        # Remove columns that are completely empty
        non_empty_cols = []
        for col in df.columns:
            # Check if column has any non-empty values
            if df[col].notna().any() and (df[col] != '').any():
                non_empty_cols.append(col)
        
        return df[non_empty_cols]

    def _remove_formatting_columns(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Remove formatting columns that are only used for Excel formatting.
        
        Args:
            df: DataFrame to clean
            
        Returns:
            DataFrame with formatting columns removed
        """
        if df.empty:
            return df
        
        # Columns to remove (formatting indicators)
        formatting_columns = ['is_section_heading', 'parent', 'is_aggregate', 'statement_type']
        
        # Remove formatting columns if they exist
        columns_to_keep = [col for col in df.columns if col not in formatting_columns]
        
        return df[columns_to_keep]

    def _merge_line_items_by_date(self, all_data_points: List[Dict], progress_callback=None) -> List[Dict]:
        """
        Merge data points from different XBRL tags that represent the same line item but have data for different dates.
        
        Args:
            all_data_points: List of data points with date, concept, and value
            progress_callback: Optional progress callback function
            
        Returns:
            List[Dict]: Merged data points with consolidated line items
        """
        def progress(msg):
            if progress_callback:
                progress_callback(msg)
            else:
                print(msg)
        
        progress("    • Merging line items with same name but different dates...")
        
        # Group data points by user-friendly title
        title_groups = {}
        for point in all_data_points:
            concept = point['concept']
            # Get user-friendly title for the concept
            friendly_title = self._get_user_friendly_title(concept)
            
            if friendly_title not in title_groups:
                title_groups[friendly_title] = []
            title_groups[friendly_title].append(point)
        
        # Merge groups that have the same title but different dates
        merged_points = []
        merge_stats = {
            'total_groups': len(title_groups),
            'merged_groups': 0,
            'total_merged_points': 0,
            'conflicts_resolved': 0
        }
        
        for friendly_title, points in title_groups.items():
            if len(points) == 1:
                # Single point, no merging needed
                merged_points.append(points[0])
                continue
            
            # Check if these points have different dates (indicating they should be merged)
            dates = [point['date'] for point in points]
            unique_dates = set(dates)
            
            if len(unique_dates) == len(points):
                # All points have different dates - this is what we want to merge
                merge_stats['merged_groups'] += 1
                merge_stats['total_merged_points'] += len(points)
                
                # Create a single merged point for each unique date
                date_value_map = {}
                for point in points:
                    date = point['date']
                    value = point['value']
                    # If multiple values for same date, take the first one (could be enhanced with validation)
                    if date not in date_value_map:
                        date_value_map[date] = value
                
                # Create merged points for each date
                for date, value in date_value_map.items():
                    merged_point = {
                        'date': date,
                        'concept': friendly_title,  # Use the friendly title as the concept
                        'value': value,
                        'merged_from': [p['concept'] for p in points]  # Track original concepts
                    }
                    merged_points.append(merged_point)
                
                progress(f"      ✓ Merged {len(points)} concepts into '{friendly_title}' ({len(date_value_map)} dates)")
            else:
                # Points have overlapping dates - resolve conflicts intelligently
                merge_stats['conflicts_resolved'] += 1
                resolved_points = self._resolve_date_conflicts(points, friendly_title, progress)
                merged_points.extend(resolved_points)
        
        progress(f"    • Merge complete: {merge_stats['merged_groups']}/{merge_stats['total_groups']} groups merged")
        progress(f"    • Conflicts resolved: {merge_stats['conflicts_resolved']}")
        progress(f"    • Total data points after merging: {len(merged_points)}")
        
        return merged_points

    def _resolve_date_conflicts(self, points: List[Dict], friendly_title: str, progress_callback=None) -> List[Dict]:
        """
        Resolve conflicts when multiple data points exist for the same date.
        Uses intelligent rules to choose the best value.
        
        Args:
            points: List of data points with potential date conflicts
            friendly_title: User-friendly title for the line item
            progress_callback: Optional progress callback function
            
        Returns:
            List[Dict]: Resolved data points
        """
        def progress(msg):
            if progress_callback:
                progress_callback(msg)
            else:
                print(msg)
        
        # Group by date
        date_groups = {}
        for point in points:
            date = point['date']
            if date not in date_groups:
                date_groups[date] = []
            date_groups[date].append(point)
        
        resolved_points = []
        
        for date, date_points in date_groups.items():
            if len(date_points) == 1:
                # No conflict for this date
                resolved_points.append(date_points[0])
                continue
            
            # Multiple points for same date - resolve conflict
            progress(f"        ⚠ Resolving conflict for '{friendly_title}' on {date} ({len(date_points)} values)")
            
            # Strategy 1: Prefer US-GAAP over non-US-GAAP
            us_gaap_points = [p for p in date_points if str(p['concept']).startswith('us-gaap:')]
            if len(us_gaap_points) == 1:
                resolved_points.append(us_gaap_points[0])
                progress(f"          ✓ Selected US-GAAP value: {us_gaap_points[0]['concept']}")
                continue
            elif len(us_gaap_points) > 1:
                # Multiple US-GAAP values - use the most specific one
                best_point = self._select_most_specific_concept(us_gaap_points)
                resolved_points.append(best_point)
                progress(f"          ✓ Selected most specific US-GAAP: {best_point['concept']}")
                continue
            
            # Strategy 2: If no US-GAAP, select the most specific concept
            best_point = self._select_most_specific_concept(date_points)
            resolved_points.append(best_point)
            progress(f"          ✓ Selected most specific concept: {best_point['concept']}")
        
        return resolved_points

    def _select_most_specific_concept(self, points: List[Dict]) -> Dict:
        """
        Select the most specific concept from a list of data points.
        More specific concepts are preferred over general ones.
        
        Args:
            points: List of data points to choose from
            
        Returns:
            Dict: The selected data point
        """
        # Define concept specificity hierarchy
        specificity_scores = {
            # Most specific (highest priority)
            'RevenueFromContractWithCustomerExcludingAssessedTax': 100,
            'SalesRevenueNet': 95,
            'Revenues': 90,
            'RevenueFromContractWithCustomer': 85,
            
            # Cost of revenue
            'CostOfRevenue': 90,
            'CostOfGoodsAndServicesSold': 85,
            'CostOfGoodsSold': 80,
            
            # Operating expenses
            'ResearchAndDevelopmentExpense': 90,
            'SellingGeneralAndAdministrativeExpense': 90,
            'OperatingExpenses': 80,
            
            # Net income
            'NetIncomeLoss': 90,
            'NetIncomeLossAvailableToCommonStockholdersBasic': 95,
            
            # Balance sheet items
            'CashAndCashEquivalentsAtCarryingValue': 90,
            'CashAndCashEquivalents': 85,
            'Cash': 80,
            
            'AccountsReceivableNet': 90,
            'AccountsReceivable': 85,
            
            'InventoryNet': 90,
            'Inventory': 85,
            
            # Default score for unknown concepts
            'default': 50
        }
        
        best_point = points[0]  # Default to first point
        best_score = specificity_scores.get('default', 50)
        
        for point in points:
            concept = point['concept']
            # Extract the concept name without namespace
            concept_name = concept.split(':')[-1] if ':' in concept else concept
            
            # Get specificity score
            score = specificity_scores.get(concept_name, specificity_scores['default'])
            
            # Boost score for US-GAAP concepts
            if concept.startswith('us-gaap:'):
                score += 10
            
            if score > best_score:
                best_score = score
                best_point = point
        
        return best_point

    def apply_text_formatting_to_titles(self, df: pd.DataFrame, formatting_style: str = 'markers') -> pd.DataFrame:
        """
        Apply text formatting directly to line item titles using formatting markers.
        
        Args:
            df: DataFrame with formatting columns (is_section_heading, parent, is_aggregate)
            formatting_style: Style of formatting to apply ('markers', 'html', 'rich_text')
            
        Returns:
            DataFrame with formatted line item titles
        """
        if df.empty:
            return df
        
        # Create a copy to avoid modifying the original
        formatted_df = df.copy()
        
        if 'Line Item' not in formatted_df.columns:
            return formatted_df
        
        # Apply formatting based on the style
        if formatting_style == 'markers':
            formatted_df['Line Item'] = formatted_df.apply(self._apply_marker_formatting, axis=1)
        elif formatting_style == 'html':
            formatted_df['Line Item'] = formatted_df.apply(self._apply_html_formatting, axis=1)
        elif formatting_style == 'rich_text':
            formatted_df['Line Item'] = formatted_df.apply(self._apply_rich_text_formatting, axis=1)
        elif formatting_style == 'unicode':
            formatted_df['Line Item'] = formatted_df.apply(self._apply_unicode_formatting, axis=1)
        
        return formatted_df
    
    def _apply_marker_formatting(self, row) -> str:
        """Apply formatting using simple text markers."""
        line_item = str(row.get('Line Item', ''))
        
        # Check formatting flags
        is_section_heading = row.get('is_section_heading', False)
        is_aggregate = row.get('is_aggregate', False)
        has_parent = row.get('parent') is not None and str(row.get('parent', '')).strip() != ''
        
        # Apply formatting markers
        if is_section_heading:
            return f"**{line_item}**"  # Bold marker
        elif is_aggregate:
            if has_parent:
                return f"  *{line_item}*"  # Italic with indentation
            else:
                return f"*{line_item}*"  # Italic
        elif has_parent:
            return f"  {line_item}"  # Indentation only
        else:
            return line_item  # No formatting
    
    def _apply_html_formatting(self, row) -> str:
        """Apply formatting using HTML tags."""
        line_item = str(row.get('Line Item', ''))
        
        # Check formatting flags
        is_section_heading = row.get('is_section_heading', False)
        is_aggregate = row.get('is_aggregate', False)
        has_parent = row.get('parent') is not None and str(row.get('parent', '')).strip() != ''
        
        # Apply HTML formatting
        if is_section_heading:
            return f"<b>{line_item}</b>"  # Bold
        elif is_aggregate:
            if has_parent:
                return f"&nbsp;&nbsp;<i>{line_item}</i>"  # Italic with indentation
            else:
                return f"<i>{line_item}</i>"  # Italic
        elif has_parent:
            return f"&nbsp;&nbsp;{line_item}"  # Indentation only
        else:
            return line_item  # No formatting
    
    def _apply_rich_text_formatting(self, row) -> str:
        """Apply formatting using rich text markers."""
        line_item = str(row.get('Line Item', ''))
        
        # Check formatting flags
        is_section_heading = row.get('is_section_heading', False)
        is_aggregate = row.get('is_aggregate', False)
        has_parent = row.get('parent') is not None and str(row.get('parent', '')).strip() != ''
        
        # Apply rich text formatting
        if is_section_heading:
            return f"[BOLD]{line_item}[/BOLD]"  # Bold
        elif is_aggregate:
            if has_parent:
                return f"  [ITALIC]{line_item}[/ITALIC]"  # Italic with indentation
            else:
                return f"[ITALIC]{line_item}[/ITALIC]"  # Italic
        elif has_parent:
            return f"  {line_item}"  # Indentation only
        else:
            return line_item  # No formatting
    
    def _apply_unicode_formatting(self, row) -> str:
        """Apply formatting using Unicode characters for visual distinction."""
        line_item = str(row.get('Line Item', ''))
        
        # Check formatting flags
        is_section_heading = row.get('is_section_heading', False)
        is_aggregate = row.get('is_aggregate', False)
        has_parent = row.get('parent') is not None and str(row.get('parent', '')).strip() != ''
        
        # Apply Unicode formatting
        if is_section_heading:
            return f"🔹 {line_item.upper()}"  # Bold section heading with bullet
        elif is_aggregate:
            if has_parent:
                return f"  📊 {line_item}"  # Aggregate with indentation and chart icon
            else:
                return f"📊 {line_item}"  # Aggregate with chart icon
        elif has_parent:
            return f"  ➤ {line_item}"  # Child item with arrow and indentation
        else:
            return f"• {line_item}"  # Regular item with bullet
    
    def export_to_excel_with_formatted_titles(self, financial_model: Dict[str, pd.DataFrame], 
                                            sensitivity_model: Dict[str, pd.DataFrame], 
                                            ticker: str, filename: str = None, 
                                            formatting_style: str = 'markers',
                                            schmoove_mode: bool = False, 
                                            progress_callback=None) -> str:
        """
        Export financial models to Excel with formatted line item titles.
        This is an alternative to the cell-level formatting approach.
        
        Args:
            financial_model: Dictionary of financial DataFrames
            sensitivity_model: Dictionary of sensitivity analysis DataFrames
            ticker: Stock ticker symbol
            filename: Output filename
            formatting_style: Style of text formatting ('markers', 'html', 'rich_text', 'unicode')
            schmoove_mode: Enable parallel processing
            progress_callback: Progress callback function
            
        Returns:
            Path to the created Excel file
        """
        def progress(msg):
            if progress_callback:
                progress_callback(msg)
            else:
                print(msg)
        
        import pandas as pd
        import os
        from datetime import datetime
        
        progress("    • Initializing Excel export with formatted titles...")
        
        # Ensure the Storage directory exists
        storage_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))), 'Storage')
        os.makedirs(storage_dir, exist_ok=True)
        
        if filename is None:
            filename = f"{ticker}_financial_model_formatted_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(storage_dir, filename)
        
        progress(f"    • Creating Excel file with {formatting_style} formatting: {filename}")
        
        # Write all DataFrames to Excel using pandas
        progress("    • Writing financial data with formatted titles...")
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            # Annual sheet
            annual_income = financial_model.get('annual_income_statement', pd.DataFrame())
            annual_balance = financial_model.get('annual_balance_sheet', pd.DataFrame())
            annual_cash_flow = financial_model.get('annual_cash_flow', pd.DataFrame())
            if not annual_income.empty or not annual_balance.empty or not annual_cash_flow.empty:
                progress("    • Writing annual financial statements with formatted titles...")
                stacked_annual = self._create_vertically_stacked_statement(annual_income, annual_balance, annual_cash_flow, period_type="Annual")
                # Apply text formatting to titles
                stacked_annual = self.apply_text_formatting_to_titles(stacked_annual, formatting_style)
                # Remove empty columns before writing to Excel
                stacked_annual = self._remove_empty_columns(stacked_annual)
                stacked_annual.to_excel(writer, sheet_name='Annual Financial Statements', index=False)
            
            # Quarterly sheet
            quarterly_income = financial_model.get('quarterly_income_statement', pd.DataFrame())
            quarterly_balance = financial_model.get('quarterly_balance_sheet', pd.DataFrame())
            quarterly_cash_flow = financial_model.get('quarterly_cash_flow', pd.DataFrame())
            if not quarterly_income.empty or not quarterly_balance.empty or not quarterly_cash_flow.empty:
                progress("    • Writing quarterly financial statements with formatted titles...")
                stacked_quarterly = self._create_vertically_stacked_statement(quarterly_income, quarterly_balance, quarterly_cash_flow, period_type="Quarterly")
                # Apply text formatting to titles
                stacked_quarterly = self.apply_text_formatting_to_titles(stacked_quarterly, formatting_style)
                # Remove empty columns before writing to Excel
                stacked_quarterly = self._remove_empty_columns(stacked_quarterly)
                stacked_quarterly.to_excel(writer, sheet_name='Quarterly Financial Statements', index=False)
            
            # Sensitivity and summary sheets
            progress("    • Writing sensitivity analysis and summary sheets...")
            for sheet_name, df in sensitivity_model.items():
                if not df.empty:
                    # Remove empty columns before writing to Excel
                    cleaned_df = self._remove_empty_columns(df)
                    cleaned_df.to_excel(writer, sheet_name=sheet_name.replace('_', ' ').title())
            
            # Summary sheet
            summary_data = {
                'Metric': ['Ticker', 'Model Created', 'Data Points', 'Scenarios Analyzed', 'Formatting Style'],
                'Value': [
                    ticker,
                    datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    sum(len(df) for df in financial_model.values() if not df.empty),
                    len(sensitivity_model.get('case_summary', pd.DataFrame())),
                    formatting_style
                ]
            }
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
        
        progress(f"    ✓ Excel file with formatted titles saved: {filepath}")
        return filepath

    def _enforce_income_statement_structure(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Ensure the DataFrame contains all required income statement line items and structure.
        Adds missing rows, enforces order, and adds section/parent info for formatting.
        """
        import numpy as np
        # Define required line items and their order
        required_order = [
            # Income Statement Section Heading
            'Income Statement',  # Section heading
            
            # Revenues Subheading
            'Revenues',  # Subheading title
            'Revenue Stream One',  # Child of Revenues
            'Revenue Stream Two',  # Child of Revenues
            'Total Revenues',  # Subtotal of Revenues
            
            # OpEx Subheading
            'OpEx',  # Subheading title
            'Cost of Goods Sold',  # Child of OpEx
            'SG&A',  # Child of OpEx
            'Misc',  # Child of OpEx
            'Total OpEx',  # Subtotal of OpEx
            
            # EBITDA (Total Revenues - Total OpEx)
            'EBITDA',  # Total Revenues - Total OpEx
            
            # Other expenses
            'Depreciation and Amortization',
            'Interest Expense',
            'Tax Expense',
            
            # Net Income (EBITDA - Depreciation and Amortization - Interest Expense - Tax Expense)
            'Net Income',  # EBITDA - Depreciation and Amortization - Interest Expense - Tax Expense
        ]
        # Parent/section mapping for indentation/formatting
        parent_map = {
            'Revenue Stream One': 'Revenues',
            'Revenue Stream Two': 'Revenues',
            'Total Revenues': 'Revenues',
            'Cost of Goods Sold': 'OpEx',
            'SG&A': 'OpEx',
            'Misc': 'OpEx',
            'Total OpEx': 'OpEx',
            'Revenues': None,
            'OpEx': None,
            'EBITDA': None,
            'Depreciation and Amortization': None,
            'Interest Expense': None,
            'Tax Expense': None,
            'Net Income': None,
            'Income Statement': None
        }
        # Create a new DataFrame with all required rows
        new_rows = []
        for item in required_order:
            if item in df.index:
                row = df.loc[item].copy()
            else:
                # Fill missing with NaN or 0
                row = pd.Series({col: np.nan for col in df.columns})
            row['Section'] = parent_map.get(item) or item
            row['Parent'] = parent_map.get(item)
            new_rows.append((item, row))
        # Build the new DataFrame
        new_df = pd.DataFrame({k: v for k, v in new_rows}).T
        new_df.index.name = 'Line Item'
        return new_df

    # --- PATCH create_financial_model to enforce structure before returning ---
    # (Find both SEC API and XBRL fallback return points)

# Example usage and testing
if __name__ == "__main__":
    sourcer = SECFileSourcer()
    
    # Show available configurations
    sourcer.print_quarter_configurations()
    
    # Use command-line arguments if provided, otherwise default to AAPL
    if len(sys.argv) > 1:
        tickers = [arg.upper() for arg in sys.argv[1:]]
    else:
        tickers = ["AAPL"]  # Default to Apple for testing

    # Get quarter configurations
    configs = sourcer.get_quarter_configurations()
    
    # Use medium_term (8 quarters) as default, but users can modify this
    quarters_to_use = configs["medium_term"]["quarters"]
    
    print(f"\nUsing {quarters_to_use} quarters of data ({quarters_to_use//4} years)")

    for ticker in tickers:
        print(f"\n{'='*40}\nTesting for ticker: {ticker}\n{'='*40}")
        print(f"Finding SEC filings for {ticker}...")
        filings = sourcer.find_sec_filings(ticker)
        print(f"Found {len(filings)} filings")
        if not filings.empty:
            print("Recent filings:")
            print(filings.head())
        print(f"\nCreating financial model for {ticker} with {quarters_to_use} quarters of data...")
        financial_model = sourcer.create_financial_model(ticker, quarters=quarters_to_use)
        if any(not df.empty for df in financial_model.values()):
            print(f"\nCreating sensitivity analysis for {ticker}...")
            sensitivity_model = sourcer.create_sensitivity_model(financial_model, ticker, quarters=quarters_to_use)
            excel_file = sourcer.export_to_excel(financial_model, sensitivity_model, ticker)
            print(f"\nModel creation complete! Check the Excel file: {excel_file}")
        else:
            print("\nNo financial data was successfully pulled. Excel file will not be created.")